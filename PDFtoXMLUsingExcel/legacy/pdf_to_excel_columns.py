import os
import subprocess
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import statistics
import argparse
import re
import statistics


# -------------------------------------------------------------
# Script Detection Configuration (Superscripts/Subscripts)
# -------------------------------------------------------------
# These thresholds detect tiny fragments (superscripts/subscripts) that should
# be merged with adjacent larger text. Very strict criteria to avoid false
# positives like drop caps or large first letters.

SCRIPT_MAX_WIDTH = 15               # Max width for scripts (drop caps are ~30-50px)
SCRIPT_MAX_HEIGHT = 14              # Max height for scripts (drop caps are ~36-48px, increased from 12 to catch 13px subscripts)
SCRIPT_MAX_TEXT_LENGTH = 3          # Max characters (usually 1-3)
SCRIPT_MAX_HORIZONTAL_GAP = 5       # Must be within 5px horizontally
SUPERSCRIPT_MIN_TOP_DIFF = -3       # Superscript can be 3px above parent
SUPERSCRIPT_MAX_TOP_DIFF = 3        # Or 3px below parent top
SUBSCRIPT_MIN_TOP_DIFF = 3          # Subscript is 3-10px below parent
SUBSCRIPT_MAX_TOP_DIFF = 10         # Maximum 10px below
SCRIPT_MAX_HEIGHT_RATIO = 0.75      # Script must be <75% of parent height

# Symbols to exclude from script detection (avoid false positives)
EXCLUDE_SYMBOLS = {'°', '™', '®', '©', '•', '·', '◦', '▪', '½', '¼', '¾', '⅓', '→', '←', '↑', '↓', '…', '‥'}


# -------------------------------------------------------------
# Script Detection Functions (Phase 1)
# -------------------------------------------------------------

def is_script_size(fragment):
    """
    Check if fragment meets size criteria for being a superscript/subscript.
    Very strict to avoid detecting drop caps or large first letters.
    """
    if fragment["width"] >= SCRIPT_MAX_WIDTH:
        return False
    if fragment["height"] >= SCRIPT_MAX_HEIGHT:
        return False
    
    text = fragment.get("text", "").strip()
    if len(text) > SCRIPT_MAX_TEXT_LENGTH:
        return False
    if not text:
        return False
    
    return True


def is_excluded_symbol(text):
    """Check if text is a symbol that should not be treated as script."""
    text = text.strip()
    
    if text in EXCLUDE_SYMBOLS:
        return True
    
    # Only allow alphanumeric scripts (excludes most symbols)
    if not text.replace('^', '').replace('_', '').isalnum():
        return True
    
    return False


def find_adjacent_parent(script_fragment, all_fragments, script_index):
    """
    Find the parent fragment for a potential superscript/subscript.
    
    Parent must be:
    - Larger in height than script
    - Adjacent horizontally (within 5px)
    - Close vertically - check if script overlaps with parent's baseline
    
    NEW: Check if top/bottom of script overlaps with baseline of other fragments
    
    Returns (parent_index, parent_fragment) or None.
    """
    script_left = script_fragment["left"]
    script_right = script_left + script_fragment["width"]
    script_top = script_fragment["top"]
    script_bottom = script_top + script_fragment["height"]
    script_height = script_fragment["height"]
    script_baseline = script_fragment["baseline"]
    
    candidates = []
    
    for i, other in enumerate(all_fragments):
        if i == script_index:
            continue
        
        # Must be larger than script
        if other["height"] <= script_height:
            continue
        
        # Script must be significantly smaller (height ratio check)
        height_ratio = script_height / other["height"]
        if height_ratio >= SCRIPT_MAX_HEIGHT_RATIO:
            continue
        
        # Check horizontal adjacency
        other_left = other["left"]
        other_right = other_left + other["width"]
        other_top = other["top"]
        other_bottom = other_top + other["height"]
        other_baseline = other["baseline"]
        
        # Is script to the right of other? (most common)
        gap_right = script_left - other_right
        if 0 <= gap_right <= SCRIPT_MAX_HORIZONTAL_GAP:
            # NEW: Check if script overlaps with parent's baseline
            # For superscripts: check if bottom of script overlaps with baseline of parent
            # For subscripts: check if top of script overlaps with baseline of parent
            
            # Allow TOP-based detection (original logic)
            top_diff = abs(script_top - other_top)
            
            # NEW: Also check baseline overlap
            # Superscript: bottom of script should be near or above the baseline of parent
            # Subscript: top of script should be near or below the baseline of parent
            baseline_overlap = False
            if script_bottom >= other_baseline - 3 and script_top <= other_baseline + 3:
                # Script overlaps with parent's baseline region
                baseline_overlap = True
            
            if top_diff <= SUBSCRIPT_MAX_TOP_DIFF or baseline_overlap:
                candidates.append((i, other, gap_right, top_diff))
        
        # Is script to the left of other? (rare)
        gap_left = other_left - script_right
        if 0 <= gap_left <= SCRIPT_MAX_HORIZONTAL_GAP:
            # Check vertical proximity using TOP and baseline overlap
            top_diff = abs(script_top - other_top)
            
            # Check baseline overlap
            baseline_overlap = False
            if script_bottom >= other_baseline - 3 and script_top <= other_baseline + 3:
                baseline_overlap = True
            
            if top_diff <= SUBSCRIPT_MAX_TOP_DIFF or baseline_overlap:
                candidates.append((i, other, gap_left, top_diff))
    
    if not candidates:
        return None
    
    # Choose closest candidate (smallest horizontal gap, then smallest vertical gap)
    candidates.sort(key=lambda x: (x[2], x[3]))
    
    parent_idx, parent, _, _ = candidates[0]
    return (parent_idx, parent)


def detect_script_type(script_fragment, parent_fragment):
    """
    Determine if script is superscript or subscript using TOP position and baseline overlap.
    
    Key insight: Check both TOP position and baseline overlap.
    - Superscript: top of script near or above parent's top
    - Subscript: top of script below parent's top
    
    NEW: Also check baseline overlap for better detection
    
    Returns "superscript", "subscript", or None.
    """
    # Calculate TOP difference (not baseline!)
    top_diff = script_fragment["top"] - parent_fragment["top"]
    
    # Calculate baseline positions
    script_baseline = script_fragment["baseline"]
    script_top = script_fragment["top"]
    script_bottom = script_top + script_fragment["height"]
    parent_baseline = parent_fragment["baseline"]
    parent_top = parent_fragment["top"]
    
    # NEW: Check baseline overlap for better classification
    # If bottom of script is at or above parent baseline → likely superscript
    # If top of script is at or below parent baseline → likely subscript
    
    # Superscript detection:
    # 1. Original: within ±3px of parent top (but not >= 3px which is subscript)
    # 2. NEW: bottom of script is above or at parent baseline
    is_superscript_by_top = (SUPERSCRIPT_MIN_TOP_DIFF <= top_diff < SUBSCRIPT_MIN_TOP_DIFF)
    is_superscript_by_baseline = (script_bottom <= parent_baseline + 2)  # Allow 2px tolerance
    
    if is_superscript_by_top or (top_diff < SUBSCRIPT_MIN_TOP_DIFF and is_superscript_by_baseline):
        return "superscript"
    
    # Subscript detection:
    # 1. Original: 3-10px below parent top
    # 2. NEW: top of script is at or below parent baseline
    is_subscript_by_top = (SUBSCRIPT_MIN_TOP_DIFF <= top_diff <= SUBSCRIPT_MAX_TOP_DIFF)
    is_subscript_by_baseline = (script_top >= parent_baseline - 2)  # Allow 2px tolerance
    
    if is_subscript_by_top or (top_diff >= SUBSCRIPT_MIN_TOP_DIFF and is_subscript_by_baseline):
        return "subscript"
    
    return None


def detect_and_mark_scripts(fragments):
    """
    Phase 1: Detect and mark superscripts/subscripts using TOP position.
    
    This modifies fragments in-place by adding:
    - is_script: bool
    - script_type: "superscript" or "subscript"
    - script_parent_idx: index of parent fragment
    
    IMPORTANT: This does NOT change grouping logic!
    Baseline grouping remains unchanged, preserving drop caps and large letters.
    """
    # Add original index to each fragment
    for i, f in enumerate(fragments):
        f["original_idx"] = i
    
    # Detect scripts
    script_count = 0
    for i, f in enumerate(fragments):
        # Default: not a script
        f["is_script"] = False
        f["script_type"] = None
        f["script_parent_idx"] = None
        
        # Check size criteria
        if not is_script_size(f):
            continue
        
        # Check if excluded symbol
        text = f.get("text", "").strip()
        if is_excluded_symbol(text):
            continue
        
        # Find adjacent parent fragment
        parent_result = find_adjacent_parent(f, fragments, i)
        if not parent_result:
            continue
        
        parent_idx, parent = parent_result
        
        # Determine script type using TOP position (not baseline!)
        script_type = detect_script_type(f, parent)
        if not script_type:
            continue
        
        # Mark as script
        f["is_script"] = True
        f["script_type"] = script_type
        f["script_parent_idx"] = parent_idx
        script_count += 1
    
    return script_count


def merge_script_with_parent(parent, scripts):
    """
    Merge one or more scripts with their parent fragment.
    
    NOW TRACKS ORIGINAL FRAGMENTS for RittDocDTD-compliant output:
    - Stores original_fragments list including parent and scripts
    - Preserves script_type metadata for proper inline element generation
    
    Args:
        parent: Parent fragment
        scripts: List of script fragments to merge (sorted by position)
    
    Returns:
        Merged fragment
    """
    merged = dict(parent)  # Copy parent
    
    # Sort scripts by left position
    scripts = sorted(scripts, key=lambda s: s["left"])
    
    # NEW: Initialize fragment tracking
    if "original_fragments" in parent:
        # Parent already has tracking from previous merge
        merged["original_fragments"] = parent["original_fragments"].copy()
    else:
        # Start tracking with parent
        parent_copy = dict(parent)
        parent_copy.pop("original_fragments", None)
        merged["original_fragments"] = [parent_copy]
    
    # Merge text with XML tags for superscripts/subscripts
    merged_text = parent["text"]
    for script in scripts:
        script_text = script["text"]

        if script["script_type"] == "superscript":
            # Use XML superscript tag: 10<superscript>7</superscript>
            merged_text += f"<superscript>{script_text}</superscript>"
        else:  # subscript
            # Use XML subscript tag: H<subscript>2</subscript>O
            merged_text += f"<subscript>{script_text}</subscript>"

        # NEW: Track the script fragment
        script_copy = dict(script)
        script_copy.pop("original_fragments", None)
        merged["original_fragments"].append(script_copy)
    
    merged["text"] = merged_text
    merged["norm_text"] = " ".join(merged_text.split()).lower()
    
    # Merge inner_xml if present (preserve formatting)
    if "inner_xml" in parent:
        merged["inner_xml"] = parent.get("inner_xml", "")
        for script in scripts:
            merged["inner_xml"] += script.get("inner_xml", script["text"])
    
    # Expand bounding box to include all scripts
    for script in scripts:
        script_right = script["left"] + script["width"]
        merged_right = merged["left"] + merged["width"]
        if script_right > merged_right:
            merged["width"] = script_right - merged["left"]
        
        # Adjust height if script extends beyond
        script_bottom = script["top"] + script["height"]
        merged_bottom = merged["top"] + merged["height"]
        if script_bottom > merged_bottom:
            merged["height"] = script_bottom - merged["top"]
    
    # Mark as having merged scripts
    merged["has_merged_scripts"] = True
    merged["merged_script_count"] = len(scripts)
    
    return merged


def merge_scripts_across_rows(rows, all_fragments):
    """
    Phase 3: Merge scripts with their parents across rows.
    
    After baseline grouping, find scripts marked in Phase 1 and
    merge them with their parent fragments even if in different rows.
    
    This is the key to fixing superscript/subscript merging while
    preserving correct baseline grouping for drop caps and large letters.
    
    Args:
        rows: List of rows (each row is list of fragments)
        all_fragments: All fragments (for looking up by original_idx)
    
    Returns:
        Updated rows with scripts merged
    """
    # Build index: original_idx -> fragment
    frag_by_idx = {}
    for row in rows:
        for f in row:
            orig_idx = f.get("original_idx")
            if orig_idx is not None:
                frag_by_idx[orig_idx] = f
    
    # Find all scripts and group by parent
    scripts_by_parent = {}
    script_indices = set()
    
    for row in rows:
        for f in row:
            if f.get("is_script"):
                parent_idx = f.get("script_parent_idx")
                if parent_idx is not None:
                    if parent_idx not in scripts_by_parent:
                        scripts_by_parent[parent_idx] = []
                    scripts_by_parent[parent_idx].append(f)
                    script_indices.add(f.get("original_idx"))
    
    # Merge scripts into their parents
    merged_rows = []
    
    for row in rows:
        new_row = []
        
        for f in row:
            orig_idx = f.get("original_idx")
            
            # Skip if this fragment is a script (will be merged into parent)
            if orig_idx in script_indices:
                continue
            
            # Check if this fragment is a parent with scripts to merge
            if orig_idx in scripts_by_parent:
                scripts = scripts_by_parent[orig_idx]
                merged = merge_script_with_parent(f, scripts)
                new_row.append(merged)
            else:
                new_row.append(f)
        
        if new_row:
            merged_rows.append(new_row)
    
    return merged_rows


# -------------------------------------------------------------
# pdftohtml -xml runner
# -------------------------------------------------------------
def run_pdftohtml_xml(pdf_path, out_xml_path):
    """
    Run `pdftohtml -xml` to convert the PDF into an XML that we can parse.

    If out_xml_path is None, we create a .xml next to the PDF.
    """
    if out_xml_path is None:
        base, _ = os.path.splitext(pdf_path)
        out_xml_path = base + "_pdftohtml.xml"

    cmd = [
        "pdftohtml",
        "-xml",
        "-hidden",
        "-nodrm",
        "-i",
        "-enc",
        "UTF-8",
        pdf_path,
        out_xml_path,
    ]
    print("Running pdftohtml (this may take a few minutes for large PDFs)...")
    print("Command:", " ".join(cmd))
    
    try:
        # Run with a reasonable timeout (10 minutes for very large PDFs)
        result = subprocess.run(cmd, check=True, timeout=600, capture_output=True, text=True)
        print("✓ pdftohtml completed successfully")
        return out_xml_path
    except subprocess.TimeoutExpired:
        print("ERROR: pdftohtml timed out after 10 minutes")
        raise
    except subprocess.CalledProcessError as e:
        print(f"ERROR: pdftohtml failed with exit code {e.returncode}")
        if e.stderr:
            print(f"stderr: {e.stderr}")
        raise


# -------------------------------------------------------------
# Reading-order & line-grouping helpers
# -------------------------------------------------------------
def _is_reference_page(fragments):
    """
    Detect if this is a reference page (Index, TOC, Glossary, etc.)
    
    Reference pages typically have:
    - Keywords like "index", "table of contents", "glossary" in text
    - High proportion of page numbers (dots followed by numbers)
    - Short text fragments (entries rather than paragraphs)
    
    Returns True if this appears to be a reference page.
    """
    if not fragments:
        return False
    
    # Check for reference keywords in first few fragments
    first_texts = " ".join(f.get("norm_text", "") for f in fragments[:10]).lower()
    reference_keywords = ["index", "table of contents", "contents", "glossary", "appendix"]
    has_keyword = any(keyword in first_texts for keyword in reference_keywords)
    
    # Check proportion of fragments with page numbers (dots followed by numbers)
    # e.g., "Introduction ............ 1" or "Chapter 1 . . . . . . 45"
    import re
    number_pattern = re.compile(r'\.{2,}\s*\d+|…+\s*\d+')  # Multiple dots followed by number
    fragments_with_numbers = sum(1 for f in fragments if number_pattern.search(f.get("text", "")))
    number_ratio = fragments_with_numbers / len(fragments) if fragments else 0
    
    # Check average text length (reference entries are typically short)
    texts = [f.get("text", "") for f in fragments]
    avg_length = sum(len(t) for t in texts) / len(texts) if texts else 0
    
    # Reference page if has keyword OR high number ratio with short entries
    return has_keyword or (number_ratio > 0.3 and avg_length < 100)


def _are_columns_side_by_side(fragments, positive_cols):
    """
    Detect if columns are side-by-side (parallel) or vertically stacked (sequential).
    
    Side-by-side columns overlap significantly in vertical space (e.g., typical 2-column layout).
    Vertically stacked columns have minimal overlap (e.g., Col1 ends, figure, then Col2 starts).
    
    Returns True if columns overlap > 50% in vertical space.
    """
    if len(positive_cols) < 2:
        return True
    
    # Get vertical range for each column
    col_ranges = {}
    for col_id in positive_cols:
        col_frags = [f for f in fragments if f["col_id"] == col_id]
        if col_frags:
            baselines = [f["baseline"] for f in col_frags]
            col_ranges[col_id] = (min(baselines), max(baselines))
    
    # Check overlap between consecutive columns
    cols = sorted(col_ranges.keys())
    for i in range(len(cols) - 1):
        col1_id = cols[i]
        col2_id = cols[i + 1]
        
        col1_range = col_ranges[col1_id]
        col2_range = col_ranges[col2_id]
        
        # Calculate overlap
        overlap_start = max(col1_range[0], col2_range[0])
        overlap_end = min(col1_range[1], col2_range[1])
        overlap = max(0, overlap_end - overlap_start)
        
        # Calculate average column height
        col1_height = col1_range[1] - col1_range[0]
        col2_height = col2_range[1] - col2_range[0]
        avg_height = (col1_height + col2_height) / 2
        
        # If overlap > 50% of average height, columns are side-by-side
        if overlap > avg_height * 0.5:
            return True
    
    # No significant overlap found - columns are vertically stacked
    return False


def _assign_column_by_column(fragments, positive_cols):
    """
    Assign blocks using true column-by-column reading order.

    For side-by-side columns, all content in one column should be read
    before moving to the next column. This is the standard reading order
    for multi-column layouts like academic papers, books, newspapers.

    Block assignment:
      1. Full-width content (col_id=0) at the top → Block 1
      2. Column 1 content (all paragraphs) → Block 2
      3. Column 2 content (all paragraphs) → Block 3
      4. Full-width content (col_id=0) in middle/bottom → Block 4
    """
    if not fragments:
        return

    # Separate full-width (col_id=0) from columnar content
    fullwidth_frags = [f for f in fragments if f["col_id"] == 0]
    columnar_frags = [f for f in fragments if f["col_id"] > 0]

    # If no columnar content, assign Block 1 to everything
    if not columnar_frags:
        for f in fragments:
            f["reading_order_block"] = 1
        return

    # Find the vertical boundary between header and columnar content
    columnar_baselines = [f["baseline"] for f in columnar_frags]
    columnar_top = min(columnar_baselines)

    # Classify full-width content as header (above columns) or other (at/below column start)
    header_frags = [f for f in fullwidth_frags if f["baseline"] < columnar_top]
    other_fullwidth_frags = [f for f in fullwidth_frags if f["baseline"] >= columnar_top]

    # Assign block numbers
    block_num = 1

    # Block 1: Full-width header content (if any)
    if header_frags:
        for f in header_frags:
            f["reading_order_block"] = block_num
        block_num += 1

    # Blocks for each column: read all content in column 1, then column 2, etc.
    for col_id in sorted(positive_cols):
        col_frags = [f for f in columnar_frags if f["col_id"] == col_id]
        if col_frags:
            for f in col_frags:
                f["reading_order_block"] = block_num
            block_num += 1

    # Final block: Full-width content that appears within or after columns
    # (figures, footnotes, etc.)
    if other_fullwidth_frags:
        for f in other_fullwidth_frags:
            f["reading_order_block"] = block_num


def _assign_interleaved(fragments):
    """
    Assign blocks using interleaved logic (for vertically stacked content).
    
    Block increments whenever col_id changes in vertical reading order.
    This handles cases like: Col1 → Figure → Col2 → Footnote
    """
    # Sort fragments by baseline (top to bottom)
    sorted_frags = sorted(fragments, key=lambda f: (f["baseline"], f.get("left", 0)))
    
    # Assign blocks based on col_id transitions
    block_num = 0
    prev_col_id = None
    
    for frag in sorted_frags:
        current_col_id = frag["col_id"]
        
        # Start a new block when col_id changes
        if current_col_id != prev_col_id:
            block_num += 1
            prev_col_id = current_col_id
        
        frag["reading_order_block"] = block_num


def assign_normalized_baselines(rows):
    """
    Assign normalized baseline to each fragment based on line grouping.

    The normalized baseline is the average of all fragment baselines in the same
    visual line (row). This handles font size variations within a line.

    Args:
        rows: List of rows from group_fragments_into_lines(), where each row
              is a list of fragments on the same visual line.
    """
    for row in rows:
        if not row:
            continue
        # Calculate average baseline for this row
        norm_baseline = sum(f["baseline"] for f in row) / len(row)
        # Round to 1 decimal place for cleaner grouping
        norm_baseline = round(norm_baseline, 1)
        for f in row:
            f["norm_baseline"] = norm_baseline


def assign_reading_order_blocks(fragments, rows):
    """
    Assign reading_order_block using baseline-driven algorithm.

    Algorithm:
    1. Use normalized baselines (from line grouping) for row-by-row processing
    2. Process baselines in ascending order (top to bottom)
    3. For each baseline:
       - If all fragments are ColID 0 (full-width): assign same block
         (continuous full-width content shares same block until structural change)
       - If single column content: assign same block
         (continuous single-col content shares same block until structural change)
       - If multiple columns detected: enter multi-column block mode
         - Process all of column 1 first (top to bottom within block)
         - Then process all of column 2 (top to bottom within block)
         - Multi-col block ends at: ColID 0, blank baseline, or page end
    4. Increment block only when structure type changes (fullwidth <-> single_col <-> multi_col)

    This ensures proper reading order for complex layouts while keeping
    continuous content together in the same block.
    """
    if not fragments:
        return

    # Ensure all fragments have norm_baseline
    for f in fragments:
        if "norm_baseline" not in f:
            f["norm_baseline"] = f["baseline"]

    # Get sorted unique normalized baselines
    sorted_baselines = sorted(set(f["norm_baseline"] for f in fragments))

    if not sorted_baselines:
        return

    # Track which fragments have been processed
    processed = set()
    reading_order_block = 1

    # Track the previous structure type for block increment decisions
    # None = not set, "fullwidth" = ColID 0 content, "single_col" = single column, "multi_col" = multi-column
    prev_structure = None

    baseline_idx = 0
    while baseline_idx < len(sorted_baselines):
        baseline = sorted_baselines[baseline_idx]

        # Get unprocessed fragments at this baseline
        baseline_frags = [f for f in fragments
                         if f["norm_baseline"] == baseline and id(f) not in processed]

        if not baseline_frags:
            baseline_idx += 1
            continue

        # Determine column structure at this baseline
        col_ids_at_baseline = sorted(set(f["col_id"] for f in baseline_frags if f["col_id"] is not None))

        # Check if all fragments are full-width (ColID 0)
        all_fullwidth = all(f["col_id"] == 0 for f in baseline_frags)

        # Check for multi-column (more than one distinct ColID, excluding 0)
        positive_cols = [c for c in col_ids_at_baseline if c > 0]
        is_multi_col = len(positive_cols) > 1

        if all_fullwidth:
            # CASE: Full-width content (ColID 0)
            # Increment block only if previous structure was different
            if prev_structure is not None and prev_structure != "fullwidth":
                reading_order_block += 1

            for f in baseline_frags:
                f["reading_order_block"] = reading_order_block
                processed.add(id(f))

            prev_structure = "fullwidth"
            baseline_idx += 1

        elif is_multi_col:
            # CASE: Multi-column content detected
            # Increment block if previous structure was different
            if prev_structure is not None and prev_structure != "multi_col":
                reading_order_block += 1

            # Find the end of this multi-column block
            # End conditions: baseline with only ColID 0, or page end
            multi_col_start_idx = baseline_idx
            multi_col_end_idx = len(sorted_baselines)  # Default to page end

            for check_idx in range(baseline_idx + 1, len(sorted_baselines)):
                check_baseline = sorted_baselines[check_idx]
                check_frags = [f for f in fragments
                              if f["norm_baseline"] == check_baseline and id(f) not in processed]

                if not check_frags:
                    # Empty baseline - end of multi-col block
                    multi_col_end_idx = check_idx
                    break

                check_cols = set(f["col_id"] for f in check_frags if f["col_id"] is not None)

                # End if we hit full-width only content
                if check_cols == {0}:
                    multi_col_end_idx = check_idx
                    break

            # Get all fragments in this multi-column block region
            block_baselines = sorted_baselines[multi_col_start_idx:multi_col_end_idx]
            block_frags = [f for f in fragments
                          if f["norm_baseline"] in block_baselines and id(f) not in processed]

            # Find all column IDs in this block (excluding ColID 0)
            cols_in_block = sorted(set(f["col_id"] for f in block_frags
                                      if f["col_id"] is not None and f["col_id"] > 0))

            # Process each column in order (left to right: col 1, col 2, ...)
            for col_id in cols_in_block:
                col_frags = [f for f in block_frags if f["col_id"] == col_id]
                # Sort by normalized baseline (top to bottom within column)
                col_frags_sorted = sorted(col_frags, key=lambda f: (f["norm_baseline"], f["left"]))

                for f in col_frags_sorted:
                    f["reading_order_block"] = reading_order_block
                    processed.add(id(f))

                # Increment block for next column
                if col_id != cols_in_block[-1]:  # Don't increment after last column
                    reading_order_block += 1

            # Handle any ColID 0 fragments within the multi-col region
            # (e.g., section headers that span columns)
            col0_in_block = [f for f in block_frags if f["col_id"] == 0 and id(f) not in processed]
            if col0_in_block:
                reading_order_block += 1
                for f in sorted(col0_in_block, key=lambda f: f["norm_baseline"]):
                    f["reading_order_block"] = reading_order_block
                    processed.add(id(f))

            prev_structure = "multi_col"
            # Jump to end of multi-col block
            baseline_idx = multi_col_end_idx

        else:
            # CASE: Single column content (one ColID, either 0 or a single positive col)
            # Keep same block for continuous single-column content
            # Increment block only if previous structure was different
            if prev_structure is not None and prev_structure != "single_col":
                reading_order_block += 1

            for f in baseline_frags:
                f["reading_order_block"] = reading_order_block
                processed.add(id(f))

            prev_structure = "single_col"
            baseline_idx += 1

    # Handle any remaining unprocessed fragments (shouldn't happen normally)
    remaining = [f for f in fragments if id(f) not in processed]
    if remaining:
        reading_order_block += 1
        for f in remaining:
            f["reading_order_block"] = reading_order_block


def compute_baseline_tolerance(baselines):
    """
    Compute how far apart two baselines can be and still be treated
    as the same row/line.
    
    ENHANCED: More lenient tolerance to handle font size/style variations
    within the same line (e.g., "MORIEL NESSAIVER, PH.D." with mixed fonts)
    """
    if len(baselines) < 2:
        return 4.0  # Increased from 2.0
    b_sorted = sorted(baselines)
    diffs = [
        b_sorted[i + 1] - b_sorted[i]
        for i in range(len(b_sorted) - 1)
        if b_sorted[i + 1] > b_sorted[i]
    ]
    if not diffs:
        return 4.0  # Increased from 2.0
    line_spacing = statistics.median(diffs)
    # Increased from 0.4 to 0.5 to be more lenient with baseline differences
    # This helps merge text with different font sizes on the same visual line
    tol = max(4.0, line_spacing * 0.5)  # Changed from min(2.0, ... * 0.4)
    return tol


def group_fragments_into_lines(fragments, baseline_tol):
    """
    Given a list of fragments (already sorted by baseline, left),
    group them into rows based on baseline tolerance.
    
    ENHANCED: Uses average baseline for the line to handle font size variations.
    Instead of comparing to the first fragment's baseline, we compare to the
    running average, which better handles mixed font sizes like:
    "MORIEL NESSAIVER, PH.D." where capital letters have different baseline than periods.
    """
    lines = []
    current = []
    current_baseline = None

    for f in fragments:
        b = f["baseline"]
        if current_baseline is None:
            current = [f]
            current_baseline = b
        elif abs(b - current_baseline) <= baseline_tol:
            # Add to current line
            current.append(f)
            # Update baseline to average of all fragments in line
            # This makes the grouping more robust to font size variations
            current_baseline = sum(frag["baseline"] for frag in current) / len(current)
        else:
            lines.append(current)
            current = [f]
            current_baseline = b

    if current:
        lines.append(current)
    return lines


def get_flattened_fragments(frag):
    """
    Get original fragments from a fragment, flattening any nested ones.

    This is critical for preserving subscript/superscript info when fragments
    are merged multiple times (e.g., script merge then inline merge then cross-page merge).

    Args:
        frag: A fragment that may have original_fragments

    Returns:
        List of flattened original fragments
    """
    if "original_fragments" in frag:
        # Fragment has nested original_fragments - use them directly
        return [dict(f) for f in frag["original_fragments"]]
    else:
        # Single fragment - wrap in list
        frag_copy = dict(frag)
        frag_copy.pop("original_fragments", None)
        return [frag_copy]


def should_merge_across_pages(last_frag, first_frag):
    """
    Determine if two fragments from consecutive pages should be merged.
    
    Merges if:
    1. Last fragment of previous page doesn't end with sentence-ending punctuation
    2. First fragment of next page continues the text naturally
    3. Both fragments are in similar columns (not full-width to column transition)
    
    Args:
        last_frag: Last fragment from previous page
        first_frag: First fragment from next page
    
    Returns:
        True if fragments should be merged
    """
    if not last_frag or not first_frag:
        return False
    
    last_text = last_frag.get("text", "").rstrip()
    first_text = first_frag.get("text", "").lstrip()
    
    if not last_text or not first_text:
        return False
    
    # Don't merge if last fragment ends with sentence-ending punctuation
    SENTENCE_ENDINGS = {'.', '!', '?', ':', ';'}
    if last_text[-1] in SENTENCE_ENDINGS:
        # Exception: ellipsis (...) should merge
        if not last_text.endswith('...'):
            return False
    
    # Don't merge if first fragment starts with capital letter followed by period
    # (likely a new section/chapter)
    if len(first_text) > 2 and first_text[0].isupper() and first_text[1:].strip().startswith('.'):
        return False
    
    # Don't merge if first fragment looks like a heading (all caps, short)
    if len(first_text) <= 50 and first_text.upper() == first_text and first_text.isalpha():
        return False
    
    # Check column compatibility
    # Don't merge if transitioning from full-width (col_id=0) to columnar or vice versa
    last_col = last_frag.get("col_id", 1)
    first_col = first_frag.get("col_id", 1)
    
    # Allow merging if:
    # - Both are same column
    # - Both are full-width (0)
    # - One is None (shouldn't happen but handle gracefully)
    if last_col is not None and first_col is not None:
        if last_col == 0 and first_col != 0:
            return False  # Full-width to columnar - likely new section
        if last_col != 0 and first_col == 0:
            return False  # Columnar to full-width - likely new section
    
    return True


def merge_paragraphs_across_pages(all_pages_data, page_order):
    """
    Merge paragraph continuations across page boundaries.
    
    When a paragraph continues from one page to the next (no sentence-ending
    punctuation at end of page), merge the fragments together. This fixes
    issues where "www.MRIsafety.com. " appears split with the period on the
    next page.
    
    Args:
        all_pages_data: Dictionary mapping page numbers to page data
        page_order: List of page numbers in order
    
    Returns:
        Modified all_pages_data with cross-page merges applied
    """
    if len(page_order) < 2:
        return all_pages_data
    
    print("Merging paragraphs across pages...")
    merge_count = 0
    
    for i in range(len(page_order) - 1):
        curr_page_num = page_order[i]
        next_page_num = page_order[i + 1]
        
        if curr_page_num not in all_pages_data or next_page_num not in all_pages_data:
            continue
        
        curr_page = all_pages_data[curr_page_num]
        next_page = all_pages_data[next_page_num]
        
        curr_fragments = curr_page.get("fragments", [])
        next_fragments = next_page.get("fragments", [])
        
        if not curr_fragments or not next_fragments:
            continue
        
        # Get last fragment of current page (highest baseline = bottom of page)
        last_frag = max(curr_fragments, key=lambda f: f.get("norm_baseline", f.get("baseline", 0)))

        # Get first fragment of next page (lowest baseline = top of page)
        first_frag = min(next_fragments, key=lambda f: f.get("norm_baseline", f.get("baseline", float('inf'))))
        
        # Check if they should be merged
        if should_merge_across_pages(last_frag, first_frag):
            # Merge the text
            last_text = last_frag.get("text", "")
            first_text = first_frag.get("text", "")
            
            # Check for soft hyphen at page boundary
            last_text_mod, first_text_mod, was_dehyphenated = remove_soft_hyphen(last_text, first_text)
            
            if was_dehyphenated:
                # Dehyphenated - join without space
                merged_text = last_text_mod + first_text_mod
            else:
                # Normal merge - join with space if needed
                if last_text.endswith(' ') or first_text.startswith(' '):
                    merged_text = last_text + first_text
                else:
                    merged_text = last_text + ' ' + first_text
            
            # Update last fragment of current page
            last_frag["text"] = merged_text
            last_frag["norm_text"] = " ".join(merged_text.split()).lower()

            # Merge original_fragments to preserve subscript/superscript info
            # Use get_flattened_fragments to handle nested original_fragments
            if "original_fragments" not in last_frag:
                last_frag["original_fragments"] = get_flattened_fragments(last_frag)
            last_frag["original_fragments"].extend(get_flattened_fragments(first_frag))

            # Propagate has_merged_scripts flag if first_frag had scripts
            if first_frag.get("has_merged_scripts"):
                last_frag["has_merged_scripts"] = True

            # Mark as merged across pages
            last_frag["merged_from_next_page"] = True

            # Mark first fragment of next page as merged (to skip in output)
            first_frag["merged_to_prev_page"] = True

            merge_count += 1
    
    if merge_count > 0:
        print(f"  Merged {merge_count} paragraph continuation(s) across pages")
    
    return all_pages_data


def remove_soft_hyphen(current_text, next_text):
    """
    Remove soft hyphens at line breaks and join hyphenated words.
    
    Handles cases like:
    - "compu-" + "tation" → "computation"
    - "self-" + "driving" → "self-driving" (keeps legitimate hyphen)
    - "pre-" + "process" → "preprocess" (removes line-break hyphen)
    
    Args:
        current_text: Text from current fragment (may end with hyphen)
        next_text: Text from next fragment (continuation)
    
    Returns:
        Tuple of (modified_current, modified_next, was_dehyphenated)
    """
    # Check if current ends with hyphen and next continues with lowercase
    if not current_text.endswith('-'):
        return current_text, next_text, False
    
    # Strip trailing whitespace from current (but keep the hyphen)
    current_stripped = current_text.rstrip()
    if not current_stripped.endswith('-'):
        return current_text, next_text, False
    
    # Strip leading whitespace from next
    next_stripped = next_text.lstrip()
    if not next_stripped:
        return current_text, next_text, False
    
    # Check if next starts with lowercase (indicates word continuation)
    first_char = next_stripped[0]
    if not first_char.islower():
        # Next starts with uppercase/number - likely a new sentence, keep hyphen
        return current_text, next_text, False
    
    # Check if this looks like a legitimate compound word
    # Look for word before the hyphen
    words_before = current_stripped[:-1].split()
    if words_before:
        last_word_before_hyphen = words_before[-1]
        
        # Unambiguous compound word prefixes (always keep hyphen)
        # These are rarely used as standalone words
        always_keep_hyphen = {
            'self', 'non', 'anti', 'co', 'semi', 'quasi', 'pseudo', 'neo', 'proto'
        }
        
        # Check if this is an unambiguous compound word
        if (last_word_before_hyphen.lower() in always_keep_hyphen and
            current_stripped.endswith(last_word_before_hyphen + '-')):
            return current_text, next_text, False
        
        # Also keep hyphen if word before is all uppercase (acronym)
        # e.g., "AI-" in "AI-powered"
        if last_word_before_hyphen.isupper() and len(last_word_before_hyphen) >= 2:
            return current_text, next_text, False
        
        # Note: Ambiguous prefixes like "pre-", "re-", "de-", "multi-" are treated
        # as line breaks by default, since in PDFs they're more commonly line breaks
        # than intentional compound words. For example:
        # - "pre-processing" is usually "preprocessing" (one word)
        # - "re-search" is usually "research" (one word), not "re-search" (search again)
        # - "de-termine" is usually "determine" (one word)
    
    # Remove the hyphen and join the words
    current_dehyphenated = current_stripped[:-1]  # Remove trailing hyphen
    
    # Preserve spacing structure
    trailing_space_in_current = current_text[len(current_stripped):]
    leading_space_in_next = next_text[:len(next_text) - len(next_stripped)]
    
    return current_dehyphenated + trailing_space_in_current, leading_space_in_next + next_stripped, True


def merge_inline_fragments_in_row(row, gap_tolerance=5.0, space_width=1.0):
    """
    Merge adjacent fragments on the same baseline using enhanced rules.
    
    NOW TRACKS ORIGINAL FRAGMENTS for RittDocDTD-compliant output:
    - Stores original_fragments list with all source fragments
    - Preserves font, size, and other metadata for each fragment
    - Enables output of inline elements (<phrase>, <emphasis>, etc.)
    
    NEW ENHANCEMENTS:
    - Increased gap tolerance (5.0px) to handle font/size variations
    - Special handling for punctuation marks (., , ; : ! ?)
    - Handles soft hyphens (word-hyphen-continuation across fragments)
    - More aggressive merging for text on same baseline

      Let:
        gap = next.left - (current.left + current.width)

      Phase 1) Punctuation merging:
         if next.text is ONLY punctuation (., , etc.)
         → merge with previous text (punctuation typically belongs to preceding word)

      Phase 2) Trailing space detection:
         if current.text ends with " " AND next.text does NOT start with " ",
         then if |gap| <= gap_tolerance
         → merge (space already present in current text).

      Phase 3) Inline-style split (no extra visible gap):
         if |gap| <= gap_tolerance
         → merge, regardless of next.text.

      Phase 4) Space-start continuation:
         if phases (1-3) fail AND next.text starts with " ",
         then if |gap - space_width| <= gap_tolerance
         → merge.
      
      Phase 5) Small gap merging (for font variations):
         if gap is small relative to text height (< 50% of min height)
         → merge (handles font size/style changes)

      Otherwise, start a new logical fragment.

    gap_tolerance and space_width are in the same units as left/width
    (usually PDF points).
    """

    if not row:
        return []

    # Sort left-to-right
    row = sorted(row, key=lambda f: f["left"])

    merged = []
    current = dict(row[0])  # copy so we don't mutate original

    # Track original fragments for RittDocDTD compliance
    # Use module-level get_flattened_fragments to preserve script info from nested merges
    current["original_fragments"] = get_flattened_fragments(row[0])

    for f in row[1:]:
        txt = f.get("text", "")
        current_txt = current.get("text", "")

        # Compute the horizontal gap between current and next
        base_end = current["left"] + current["width"]
        gap = f["left"] - base_end

        should_merge = False
        
        # --- SPECIAL CASE: Punctuation merging ---
        # Punctuation marks (., , ; : ! ? ) etc.) should ALWAYS merge with preceding text
        # They often appear as separate elements due to different positioning
        PUNCTUATION_CHARS = {'.', ',', ';', ':', '!', '?', ')', ']', '}', '"', "'", '…', '»', '›'}
        txt_stripped = txt.strip()
        
        # Check if next fragment is ONLY punctuation (possibly with whitespace)
        if txt_stripped and all(c in PUNCTUATION_CHARS or c.isspace() for c in txt_stripped):
            # Punctuation element - merge if reasonably close (within 10px)
            if gap <= 10.0:
                should_merge = True
        
        # --- SPECIAL CASE: Bullet point merging ---
        # Detect if current is a bullet character and next is text
        # Bullets are often positioned differently (different baseline/height)
        # So we need more lenient merging for bullets
        if not should_merge:
            BULLET_CHARS = {'•', '●', '○', '■', '□', '▪', '▫', '·', '-', '*', '–', '—', '→', '⇒', '▸', '►'}
            current_stripped = current_txt.strip()
            
            if current_stripped in BULLET_CHARS and len(current_stripped) == 1:
                # Current is a bullet character - merge with following text if reasonably close
                # Allow larger gap (up to 20px) since bullets are often positioned differently
                if gap <= 20.0:  # More lenient for bullets
                    should_merge = True

        # --- Phase 2: trailing space detection ---
        # If current ends with space and next does NOT start with space
        if not should_merge and current_txt.endswith(" ") and not txt.startswith(" "):
            # Check if gap is small (approximately zero)
            if abs(gap) <= gap_tolerance:
                should_merge = True

        # --- Phase 3: inline-style / no-gap merge ---
        if not should_merge:
            nogap = abs(gap) <= gap_tolerance
            if nogap:
                should_merge = True

        # --- Phase 4: starts-with-space + "space gap" (± tolerance) ---
        if not should_merge:
            if txt.startswith(" "):
                space_gap_ok = abs(gap - space_width) <= gap_tolerance
                if space_gap_ok:
                    should_merge = True
        
        # --- Phase 5: Small gap relative to text height (handles font variations) ---
        # If gap is small compared to text height, likely same word/phrase with font change
        if not should_merge:
            current_height = current.get("height", 0)
            next_height = f.get("height", 0)
            min_height = min(current_height, next_height) if current_height > 0 and next_height > 0 else 0
            
            # If gap is less than 50% of the smaller text height, merge
            # This handles cases like "MORIEL NESSAIVER" where fonts change mid-word
            if min_height > 0 and gap > 0 and gap < min_height * 0.5:
                should_merge = True
        
        # --- Phase 6: Continuation words (handles font style changes) ---
        # If previous text ends with continuation words, merge even with moderate gap
        # This handles cases like "including Journal Name" where font style changes
        if not should_merge and gap > 0 and gap <= 15.0:
            current_txt_lower = current_txt.lower().rstrip()
            continuation_words = {'including', 'and', 'or', 'the', 'for', 'in', 'of', 'to', 'a', 'an',
                                 'as', 'with', 'from', 'by', 'at', 'on', 'into', 'through', 'during',
                                 'such', 'both', 'each', 'all', 'other', 'these', 'those', 'many'}
            
            # Check if current text ends with a continuation word
            for word in continuation_words:
                if current_txt_lower.endswith(word) or current_txt_lower.endswith(word + ','):
                    should_merge = True
                    break

        if should_merge:
            # NEW: Check for soft hyphen before merging
            current_txt_before = current.get("text", "")
            next_txt_before = txt
            current_txt_dehyph, next_txt_dehyph, was_dehyphenated = remove_soft_hyphen(
                current_txt_before, next_txt_before
            )
            
            if was_dehyphenated:
                # Update texts after dehyphenation
                current["text"] = current_txt_dehyph + next_txt_dehyph
                # Also update inner_xml to reflect dehyphenation
                # Remove trailing hyphen from inner_xml if present
                current_inner = current.get("inner_xml", current_txt_before)
                next_inner = f.get("inner_xml", next_txt_before)
                if current_inner.rstrip().endswith('-'):
                    current_inner = current_inner.rstrip()[:-1] + current_inner[len(current_inner.rstrip()):]
                current["inner_xml"] = current_inner + next_inner
            else:
                # Merge: append text as-is (keep whatever spaces are in txt)
                current["text"] = current.get("text", "") + txt
                # Merge XML content to preserve formatting
                current["inner_xml"] = current.get("inner_xml", "") + f.get("inner_xml", txt)
            
            current["norm_text"] = " ".join(current["text"].split()).lower()

            # Expand width to cover the new fragment
            prev_end = current["left"] + current["width"]
            right = max(prev_end, f["left"] + f["width"])
            current["width"] = right - current["left"]
            
            # NEW: Track the merged fragment (flatten nested original_fragments)
            current["original_fragments"].extend(get_flattened_fragments(f))
        else:
            # Start a new logical fragment
            merged.append(current)
            current = dict(f)

            # NEW: Initialize tracking for new fragment (flatten nested original_fragments)
            current["original_fragments"] = get_flattened_fragments(f)

    merged.append(current)
    return merged




# ---------------------------------------
# Fragment filtering (headers/footers)
# ---------------------------------------
def should_skip_fragment(norm_txt, top, height, page_height, seen_footer_texts):

    # 1) Skip if outside visible page render area
    if top > page_height * 1.05:     # below page
        return True
    if top < -20:                    # above page
        return True

    # 2) Skip file names, indesign junk, timestamps
    if re.search(r"\.indd\b", norm_txt):
        return True
    if re.search(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", norm_txt):
        return True  # dates like 12/18/18
    if re.search(r"\b\d{1,2}:\d{2}\s*(am|pm)\b", norm_txt):
        return True
    if re.fullmatch(r"[a-z0-9_\-]+\s+vi|vii|iv", norm_txt):
        return True

    # 3) Skip extremely small-height invisible text
    # (common for print-layer artifacts)
    if int(height) < 6: 
        return True

    # 4) Header/footer filtering - skip repeated text at page edges
    # Check if text is in header zone (top 12%) or footer zone (bottom 12%)
    if page_height > 0:
        norm_top = top / page_height
        is_header_zone = norm_top < 0.12
        is_footer_zone = norm_top > 0.85

        if (is_header_zone or is_footer_zone) and norm_txt in seen_footer_texts:
            return True

        # 5) Skip standalone page numbers at header/footer zones
        # These are captured as page IDs, don't need them in content
        if is_header_zone or is_footer_zone:
            text_stripped = norm_txt.strip()
            # Arabic page numbers (1-9999)
            if re.match(r'^\d{1,4}$', text_stripped):
                return True
            # Roman numerals (i, ii, iii, iv, v, vi, vii, viii, ix, x, etc.)
            if re.match(r'^[ivxlcdm]+$', text_stripped, re.IGNORECASE):
                return True

    return False


# -------------------------------------------------------------
# Column detection
# -------------------------------------------------------------

def compute_dynamic_column_gap_threshold(fragments, page_width, baseline_tolerance=3.0):
    """
    Compute a dynamic column gap threshold based on typical word/fragment spacing.

    A column gap should be significantly larger than normal word gaps within text.
    This analyzes the gaps between consecutive fragments on the same baseline
    and returns a threshold that distinguishes column gaps from word gaps.

    Uses norm_baseline (if available) for more accurate line grouping.

    Handles three scenarios:
    1. Bimodal gaps: small word gaps + large column gaps → threshold between them
    2. Only large gaps (column-only): likely pure columns → use fraction of gap
    3. Only small gaps: word gaps only → multiply to get column threshold

    Args:
        fragments: List of fragment dicts with 'left', 'width', 'baseline' keys
        page_width: Width of the page in pixels
        baseline_tolerance: Tolerance for grouping fragments into same line

    Returns:
        float: The dynamic column gap threshold
    """
    if not fragments or len(fragments) < 2:
        return page_width * 0.10  # Fallback to 10% of page width

    # Use norm_baseline if available (more accurate), otherwise fall back to baseline
    def get_baseline(f):
        return f.get("norm_baseline", f["baseline"])

    # Group fragments by baseline (same horizontal line)
    baseline_groups = {}
    sorted_frags = sorted(fragments, key=get_baseline)

    current_baseline = None
    current_group = []

    for frag in sorted_frags:
        b = get_baseline(frag)
        if current_baseline is None:
            current_baseline = b
            current_group = [frag]
        elif abs(b - current_baseline) <= baseline_tolerance:
            current_group.append(frag)
            # Update baseline to running average for better grouping
            current_baseline = sum(get_baseline(f) for f in current_group) / len(current_group)
        else:
            if current_group:
                baseline_groups[current_baseline] = current_group
            current_baseline = b
            current_group = [frag]

    if current_group:
        baseline_groups[current_baseline] = current_group

    # Collect all gaps between consecutive fragments on same line
    all_gaps = []

    for baseline, frags in baseline_groups.items():
        if len(frags) < 2:
            continue

        # Sort fragments left to right
        sorted_line_frags = sorted(frags, key=lambda f: f["left"])

        # Measure gaps between consecutive fragments
        for i in range(len(sorted_line_frags) - 1):
            frag_a = sorted_line_frags[i]
            frag_b = sorted_line_frags[i + 1]

            # Gap = start of next fragment - end of current fragment
            gap = frag_b["left"] - (frag_a["left"] + frag_a["width"])

            # Only consider positive gaps (ignore overlapping fragments)
            if gap > 0:
                all_gaps.append(gap)

    if not all_gaps:
        return page_width * 0.10  # Fallback

    # Sort gaps for analysis
    gaps_sorted = sorted(all_gaps)
    min_gap = gaps_sorted[0]
    max_gap = gaps_sorted[-1]
    median_gap = statistics.median(gaps_sorted)

    # Check for bimodal distribution (word gaps + column gaps)
    # A large ratio between max and min suggests bimodal
    gap_ratio = max_gap / min_gap if min_gap > 0 else 1.0

    # Calculate 25th and 75th percentiles
    q25_idx = max(0, int(len(gaps_sorted) * 0.25))
    q75_idx = min(len(gaps_sorted) - 1, int(len(gaps_sorted) * 0.75))
    q25_gap = gaps_sorted[q25_idx]
    q75_gap = gaps_sorted[q75_idx]

    # Fix for full-line XML fragments: When XML has full text lines (not words),
    # column gaps are small but short lines create outlier large gaps that inflate variance.
    # Use IQR-based outlier filtering for wide fragments.
    avg_fragment_width = statistics.mean([f["width"] for f in fragments])
    fragments_are_wide = avg_fragment_width > page_width * 0.25

    if fragments_are_wide and len(all_gaps) >= 8:
        q1_idx = int(len(gaps_sorted) * 0.25)
        q3_idx = int(len(gaps_sorted) * 0.75)
        q1 = gaps_sorted[q1_idx]
        q3 = gaps_sorted[q3_idx]
        iqr = q3 - q1

        lower_bound = q1 - 1.5 * iqr
        upper_bound = q3 + 1.5 * iqr
        filtered_gaps = [g for g in all_gaps if lower_bound <= g <= upper_bound]

        if len(filtered_gaps) >= 5:
            filtered_mean = statistics.mean(filtered_gaps)
            filtered_stdev = statistics.stdev(filtered_gaps) if len(filtered_gaps) > 1 else 0
            filtered_cv = filtered_stdev / filtered_mean if filtered_mean > 0 else 0

            if filtered_cv < 0.5:
                filtered_min = min(filtered_gaps)
                column_gap_threshold = filtered_min * 0.70
                column_gap_threshold = max(column_gap_threshold, page_width * 0.02)
                return min(column_gap_threshold, page_width * 0.25)

    # Threshold for considering a gap "large" (potential column gap)
    # Use 3% instead of 5% to catch narrower column gaps
    large_gap_threshold = page_width * 0.03  # 3% of page width

    # Count small vs large gaps
    small_gaps = [g for g in all_gaps if g < large_gap_threshold]
    large_gaps = [g for g in all_gaps if g >= large_gap_threshold]

    if small_gaps and large_gaps and len(large_gaps) >= 3:
        # BIMODAL: We have both word gaps and column gaps
        # Threshold should be between max small gap and min large gap
        max_small = max(small_gaps)
        min_large = min(large_gaps)
        column_gap_threshold = (max_small + min_large) / 2.0
    elif large_gaps and not small_gaps:
        # ALL LARGE GAPS: Likely a multi-column layout with no word breaks
        # Use a fraction of the minimum large gap (e.g., 70% of smallest column gap)
        column_gap_threshold = min(large_gaps) * 0.70
    elif gap_ratio > 3.0 and len(all_gaps) >= 5:
        # Significant spread in gap sizes - try to find natural break point
        # Use a multiple of the 25th percentile (word gaps) but less than median
        column_gap_threshold = max(
            q25_gap * 3.0,
            (q25_gap + median_gap) / 2.0,
            page_width * 0.05
        )
    else:
        # UNIMODAL SMALL GAPS: Only word gaps present
        # Column gap should be significantly larger
        column_gap_threshold = max(
            q75_gap * 4.0,
            median_gap * 6.0,
            page_width * 0.05
        )

    # Ensure threshold is within reasonable bounds
    column_gap_threshold = max(column_gap_threshold, page_width * 0.03)  # At least 3%
    column_gap_threshold = min(column_gap_threshold, page_width * 0.25)  # At most 25%

    return column_gap_threshold


def detect_columns_baseline_based(fragments, page_width, baseline_tolerance=3.0,
                                   min_consistent_baselines=8, max_cols=4):
    """
    Detect multi-column layout by analyzing fragments on the same baseline.

    TRUE multi-column detection: Checks if fragments at DIFFERENT X positions
    appear SIDE-BY-SIDE on the SAME horizontal line. This correctly handles:
    - Indented paragraphs (different baselines for each indent level)
    - Bulleted lists (bullets and text on same line, but continuous)
    - Figure captions (different indent but not side-by-side with body text)

    Uses norm_baseline (if available) for more accurate line grouping.

    Args:
        fragments: List of fragment dicts
        page_width: Width of the page
        baseline_tolerance: Tolerance for grouping fragments into same line
        min_consistent_baselines: Minimum number of baselines that must show
                                  the same column pattern to confirm multi-column
        max_cols: Maximum number of columns to detect

    Returns:
        List of column start X positions (sorted), or single-element list if single column
    """
    if not fragments or len(fragments) < 2:
        # Return leftmost position as single column
        if fragments:
            return [min(f["left"] for f in fragments)]
        return []

    # Compute dynamic threshold based on actual content
    column_gap_threshold = compute_dynamic_column_gap_threshold(
        fragments, page_width, baseline_tolerance
    )

    # Use norm_baseline if available (more accurate), otherwise fall back to baseline
    def get_baseline(f):
        return f.get("norm_baseline", f["baseline"])

    # Group fragments by baseline (same horizontal line)
    baseline_groups = {}
    sorted_frags = sorted(fragments, key=get_baseline)

    current_baseline = None
    current_group = []

    for frag in sorted_frags:
        b = get_baseline(frag)
        if current_baseline is None:
            current_baseline = b
            current_group = [frag]
        elif abs(b - current_baseline) <= baseline_tolerance:
            current_group.append(frag)
            current_baseline = sum(get_baseline(f) for f in current_group) / len(current_group)
        else:
            if current_group:
                baseline_groups[current_baseline] = current_group
            current_baseline = b
            current_group = [frag]

    if current_group:
        baseline_groups[current_baseline] = current_group

    # For each baseline, find column starts (positions after significant gaps)
    # A column start is: the leftmost position OR position after a gap > threshold
    all_column_starts_by_baseline = []

    for baseline, frags in baseline_groups.items():
        if len(frags) < 1:
            continue

        # Sort fragments left to right
        sorted_line_frags = sorted(frags, key=lambda f: f["left"])

        # First fragment's left is always a potential column start
        line_col_starts = [sorted_line_frags[0]["left"]]

        # Find additional column starts (after significant gaps)
        for i in range(len(sorted_line_frags) - 1):
            frag_a = sorted_line_frags[i]
            frag_b = sorted_line_frags[i + 1]

            gap = frag_b["left"] - (frag_a["left"] + frag_a["width"])

            # Significant gap indicates a new column
            if gap > column_gap_threshold:
                line_col_starts.append(frag_b["left"])

        # Only record if this line has potential multi-column structure
        # (i.e., more than one column start position on same baseline)
        if len(line_col_starts) > 1:
            all_column_starts_by_baseline.append(line_col_starts)

    # If not enough baselines show multi-column pattern, it's single column
    if len(all_column_starts_by_baseline) < min_consistent_baselines:
        # Single column - return leftmost position
        leftmost = min(f["left"] for f in fragments)
        return [leftmost]

    # Cluster the column start positions across baselines
    # We need to find CONSISTENT column boundaries
    #
    # Strategy: For each baseline's column starts, the Nth start should be
    # at approximately the same X position across all baselines

    # First, determine the most common number of columns
    num_cols_counter = {}
    for col_starts in all_column_starts_by_baseline:
        n = len(col_starts)
        num_cols_counter[n] = num_cols_counter.get(n, 0) + 1

    # Find the most common column count (but respect max_cols)
    most_common_num_cols = max(num_cols_counter, key=num_cols_counter.get)
    most_common_num_cols = min(most_common_num_cols, max_cols)

    # Filter to only baselines with this number of columns
    consistent_baselines = [
        cs for cs in all_column_starts_by_baseline
        if len(cs) == most_common_num_cols
    ]

    # Need at least min_consistent_baselines with same column structure
    if len(consistent_baselines) < min_consistent_baselines:
        # Not enough consistency - might be false detection
        # Check if there's a secondary common count
        if len(num_cols_counter) > 1:
            sorted_counts = sorted(num_cols_counter.items(), key=lambda x: x[1], reverse=True)
            for num_cols, count in sorted_counts[1:]:
                if count >= min_consistent_baselines and num_cols <= max_cols:
                    consistent_baselines = [
                        cs for cs in all_column_starts_by_baseline
                        if len(cs) == num_cols
                    ]
                    most_common_num_cols = num_cols
                    break

        if len(consistent_baselines) < min_consistent_baselines:
            leftmost = min(f["left"] for f in fragments)
            return [leftmost]

    # Calculate average column start positions
    # For each column index, average the X positions across all consistent baselines
    final_col_starts = []

    for col_idx in range(most_common_num_cols):
        col_positions = [cs[col_idx] for cs in consistent_baselines if col_idx < len(cs)]
        if col_positions:
            avg_position = sum(col_positions) / len(col_positions)
            final_col_starts.append(avg_position)

    # Validate: columns should be reasonably spaced apart
    # If two "columns" are too close, merge them
    min_col_separation = page_width * 0.10  # At least 10% of page width apart

    merged_col_starts = []
    for pos in sorted(final_col_starts):
        if not merged_col_starts:
            merged_col_starts.append(pos)
        elif pos - merged_col_starts[-1] >= min_col_separation:
            merged_col_starts.append(pos)
        # else: too close to previous, skip (merge into previous)

    # Final sanity check: if only one column start remains, it's single column
    if len(merged_col_starts) <= 1:
        leftmost = min(f["left"] for f in fragments)
        return [leftmost]

    return sorted(merged_col_starts)


def detect_column_starts(fragments, page_width, max_cols=4,
                         min_cluster_size=15, min_cluster_ratio=0.10):
    """
    Detect approximate x-start of each text column using baseline-based analysis
    with fallback to left-position clustering.

    PRIMARY ALGORITHM: Analyzes fragments on the SAME BASELINE (same horizontal line)
    to detect true multi-column layouts where content runs in parallel.

    FALLBACK: If baseline detection fails (returns single column), checks for
    distinct left-position clusters. This handles cases like Table of Contents
    where columns don't have content side-by-side on the same baseline.

    Args:
        fragments: List of fragment dicts with 'left', 'width', 'baseline' keys
        page_width: Width of the page in pixels
        max_cols: Maximum number of columns to detect (default 4)
        min_cluster_size: Minimum fragments in a cluster for fallback detection
        min_cluster_ratio: Minimum ratio of fragments for fallback detection

    Returns:
        List of column start X positions (sorted).
        Single-element list if single column layout detected.
    """
    # Try baseline-based detection first
    col_starts = detect_columns_baseline_based(
        fragments,
        page_width,
        baseline_tolerance=3.0,
        min_consistent_baselines=8,
        max_cols=max_cols
    )

    # If baseline detection found multi-column, use it
    if len(col_starts) > 1:
        return col_starts

    # Fallback: Use left-position clustering for cases like TOC pages
    # where columns don't have content side-by-side on same baseline
    fallback_cols = detect_columns_by_left_clustering(
        fragments, page_width, max_cols, min_cluster_size, min_cluster_ratio
    )
    return fallback_cols


def detect_columns_by_left_clustering(fragments, page_width, max_cols=4,
                                       min_cluster_size=15, min_cluster_ratio=0.10):
    """
    Fallback column detection using left-position clustering.

    This handles pages where columns exist but don't have content side-by-side
    on the same baseline (e.g., Table of Contents, index pages).

    The algorithm:
    1. Collect all left positions
    2. Find distinct clusters of left positions
    3. Require clusters to be well-separated (> 20% of page width apart)
    4. Require each cluster to have significant fragment count

    Args:
        fragments: List of fragment dicts with 'left' key
        page_width: Width of the page
        max_cols: Maximum columns to detect
        min_cluster_size: Minimum fragments per cluster
        min_cluster_ratio: Minimum ratio of total fragments per cluster

    Returns:
        List of column start X positions (sorted), or single-element if single column
    """
    if not fragments or len(fragments) < 2:
        if fragments:
            return [min(f["left"] for f in fragments)]
        return []

    # Collect all left positions
    left_positions = [f["left"] for f in fragments]

    # Simple clustering: group left positions that are within tolerance
    cluster_tolerance = page_width * 0.05  # 5% of page width
    min_separation = page_width * 0.20  # Clusters must be 20% apart to be columns

    # Sort positions and find clusters
    sorted_lefts = sorted(left_positions)

    clusters = []
    current_cluster = [sorted_lefts[0]]

    for left in sorted_lefts[1:]:
        if left - current_cluster[-1] <= cluster_tolerance:
            current_cluster.append(left)
        else:
            clusters.append(current_cluster)
            current_cluster = [left]

    clusters.append(current_cluster)

    # Filter clusters by size
    min_size = max(min_cluster_size, len(fragments) * min_cluster_ratio)
    significant_clusters = [c for c in clusters if len(c) >= min_size]

    if len(significant_clusters) < 2:
        # Not enough significant clusters for multi-column
        return [min(left_positions)]

    # Get cluster centers (average left position)
    cluster_centers = []
    for cluster in significant_clusters:
        center = sum(cluster) / len(cluster)
        cluster_centers.append(center)

    # Sort by position
    cluster_centers.sort()

    # Merge clusters that are too close together
    merged_centers = [cluster_centers[0]]
    for center in cluster_centers[1:]:
        if center - merged_centers[-1] >= min_separation:
            merged_centers.append(center)

    # Limit to max_cols
    merged_centers = merged_centers[:max_cols]

    if len(merged_centers) < 2:
        return [min(left_positions)]

    return merged_centers


def is_vertical_spine_text(text, left, width, height,
                           page_width, page_height, rotation_deg):
    """
    Heuristic to filter out vertical spine text at the right border of the page,
    e.g. 'INTRODUCTION' printed vertically.

    We consider a fragment a vertical spine candidate if:
      - text is very short (<= 3 chars after stripping),
      - it sits near the right margin of the page,
      - and it is clearly vertical, either:
          * rotated ~90/270 degrees, OR
          * very tall & narrow.
    """
    if not text:
        return False

    t = text.strip()
    if len(t) > 3:
        return False

    # Near right border (tweak 0.8 if needed)
    if left <= page_width * 0.8:
        return False

    # Normal horizontal text: rotation 0 and not tall/narrow
    is_vertical_rot = abs(rotation_deg) in (90, 270)
    is_tall_narrow = height > width * 2.0

    if not (is_vertical_rot or is_tall_narrow):
        return False

    return True


def assign_column_ids(fragments, page_width, col_starts):
    """
    Assign a column id to each fragment on the page.

    col_id:
      1..N = one of the detected column clusters
      0    = full-width (spanning the entire page / multiple columns)
    """
    if not fragments:
        return

    # If there's effectively only one column start, treat everything as col 1
    if len(col_starts) <= 1:
        for f in fragments:
            f["col_id"] = 1
        return

     # Instead of using width/page_width, treat a fragment as "full-width"
    # only if it nearly touches BOTH left and right margins.
    margin_ratio = 0.05  # 5% of page width as margin
    left_margin  = page_width * margin_ratio
    right_margin = page_width * (1.0 - margin_ratio)

    # Calculate column boundaries (midpoints between adjacent column starts)
    # These boundaries define the territory for each column
    boundaries = []
    for i in range(len(col_starts) - 1):
        midpoint = (col_starts[i] + col_starts[i + 1]) / 2.0
        boundaries.append(midpoint)

    for f in fragments:
        left  = f["left"]
        right = f["left"] + f["width"]
        width = f["width"]

        # PRIORITY 1: Check if fragment truly spans from left margin to right margin
        # This is the only case where we should definitely assign ColID 0
        if left <= left_margin and right >= right_margin:
            f["col_id"] = 0
            continue

        # PRIORITY 1.5: Check for CENTERED text that should be treated as full-width
        # Centered text (like titles "DEDICATION") doesn't span full width but should be ColID 0
        # Criteria:
        #   - Fragment's center is close to page center (within 15% of page width)
        #   - Fragment's left position doesn't align with any column start (within 5% tolerance)
        #   - OR fragment is wide enough (>30% page width) and centered
        fragment_center = left + width / 2.0
        page_center = page_width / 2.0
        center_tolerance = page_width * 0.15  # 15% tolerance for center detection

        is_centered = abs(fragment_center - page_center) <= center_tolerance

        if is_centered:
            # Check if left position aligns with any column start
            # Use tight tolerance (5%) to avoid matching content that's clearly in a column
            col_start_tolerance = page_width * 0.05  # 5% tolerance for column alignment
            aligns_with_column = any(
                abs(left - col_start) <= col_start_tolerance
                for col_start in col_starts
            )

            # Wide centered text (>30% page width) should always be ColID 0
            # even if close to a column start - it's spanning across columns
            is_wide_centered = width >= page_width * 0.30

            # If centered and (doesn't align with column starts OR is wide), it's full-width content
            if not aligns_with_column or is_wide_centered:
                f["col_id"] = 0
                continue

        # PRIORITY 2: Check if left position clearly places fragment in a column
        # If left position is within a column's territory, assign to that column
        # regardless of width (width check should not override clear column membership)
        assigned_by_position = False

        if left < boundaries[0]:
            # Left edge is clearly in Column 1's territory
            f["col_id"] = 1
            assigned_by_position = True
        elif len(boundaries) > 1 and left >= boundaries[-1]:
            # Left edge is clearly in last column's territory
            f["col_id"] = len(col_starts)
            assigned_by_position = True
        else:
            # Find which column territory the left edge falls into
            for i in range(len(boundaries)):
                if i == len(boundaries) - 1:
                    f["col_id"] = i + 2
                    assigned_by_position = True
                    break
                elif left < boundaries[i + 1]:
                    f["col_id"] = i + 1
                    assigned_by_position = True
                    break

        # PRIORITY 3: Only apply width check if position assignment didn't happen
        # AND the fragment spans across multiple column boundaries
        if not assigned_by_position:
            # Fragment position is ambiguous - use width as fallback
            if width >= page_width * 0.45:
                f["col_id"] = 0
            else:
                f["col_id"] = 1  # Default to column 1


def reassign_misclassified_col0_fragments(fragments, page_width, col_starts):
    """
    Reassign ColID 0 fragments to the correct column if they're clearly
    within a single column's bounds and not actually full-width.

    This fixes cases where fragments were incorrectly classified as full-width (ColID 0)
    but are actually within a specific column.
    """
    if not fragments or len(col_starts) <= 1:
        return

    margin_ratio = 0.05
    left_margin = page_width * margin_ratio
    right_margin = page_width * (1.0 - margin_ratio)

    for f in fragments:
        # Only process fragments currently assigned to ColID 0
        if f["col_id"] != 0:
            continue

        left = f["left"]
        right = f["left"] + f["width"]
        width = f["width"]
        x_center = (left + right) / 2.0

        # Check if fragment actually spans full width - if so, keep ColID 0
        if left <= left_margin and right >= right_margin:
            continue  # Actually full-width, keep ColID 0

        # Check if fragment is very wide - if so, keep ColID 0
        if width >= page_width * 0.45:
            continue  # Wide enough to be full-width, keep ColID 0

        # Fragment is ColID 0 but doesn't span full width
        # Reassign to nearest column based on center position
        best = min(col_starts, key=lambda c: abs(x_center - c))
        f["col_id"] = col_starts.index(best) + 1


def group_col0_by_vertical_gap(fragments, typical_line_height, page_width=None, page_height=None):
    """
    Group ColID 0 fragments based on vertical gap - ONLY for wide fragments.

    When a fragment gets ColID 0 assigned, check the vertical gap to the next fragment.
    If the gap is small (within typical line height range) AND the next fragment is
    also wide enough to be full-width, assign ColID 0 to the next fragment.

    This ensures that multi-line titles, captions, and other full-width content
    that should logically be together all get ColID 0, WITHOUT incorrectly
    converting narrow column text to ColID 0.

    Args:
        fragments: List of fragments (must have baseline, col_id, width, left)
        typical_line_height: Typical line height for gap threshold
        page_width: Page width for determining if fragment is truly full-width
        page_height: Page height (unused, kept for API compatibility)
    """
    if not fragments:
        return

    # If we don't have page_width, skip this function to avoid incorrect ColID 0 propagation
    if page_width is None:
        return

    # Sort by baseline (top to bottom)
    sorted_frags = sorted(fragments, key=lambda f: f["baseline"])

    # Maximum vertical gap to consider fragments as part of the same ColID 0 group
    # Use 1.5x typical line height as threshold (same as paragraph detection)
    max_gap = typical_line_height * 1.5

    # Width thresholds for considering a fragment as "wide enough" for ColID 0
    # A fragment must be at least 40% of page width to be grouped as ColID 0
    min_width_ratio = 0.40
    min_width_for_col0 = page_width * min_width_ratio

    i = 0
    while i < len(sorted_frags):
        current = sorted_frags[i]

        # Only process if current fragment has ColID 0
        if current["col_id"] != 0:
            i += 1
            continue

        # Current fragment has ColID 0, check subsequent fragments
        j = i + 1
        while j < len(sorted_frags):
            next_frag = sorted_frags[j]

            # Calculate vertical gap
            current_bottom = current["top"] + current["height"]
            next_top = next_frag["top"]
            vertical_gap = next_top - current_bottom

            # If gap is too large, stop grouping
            if vertical_gap > max_gap:
                break

            # Check if we should propagate ColID 0 to next fragment
            next_width = next_frag.get("width", 0)

            if next_frag["col_id"] != 0:
                # Only propagate ColID 0 if fragment is wide enough (>= 40% page width)
                # This groups multi-line titles/captions while preserving column text
                if next_width >= min_width_for_col0:
                    next_frag["col_id"] = 0
                    current = next_frag
                else:
                    # Fragment is too narrow (within a column), stop propagating
                    break
            elif next_frag["col_id"] == 0:
                # Already ColID 0, continue
                current = next_frag
            else:
                break

            j += 1

        # Move to the next ungrouped fragment
        i = j if j > i + 1 else i + 1


def maintain_col0_within_baseline(fragments, baseline_tol, max_word_gap=50):
    """
    Maintain ColID 0 for horizontally adjacent fragments on the same baseline.

    If a fragment on a baseline has ColID 0, only propagate ColID 0 to fragments
    that are horizontally adjacent (within max_word_gap). This prevents ColID 0
    from spreading across column boundaries to unrelated text.

    This fixes cases where the last small fragment on a ColID 0 line gets incorrectly
    assigned to ColID 1, while NOT affecting fragments in different column territories.

    Args:
        fragments: List of fragments with baseline, left, width, and col_id
        baseline_tol: Tolerance for grouping fragments into same baseline
        max_word_gap: Maximum horizontal gap to consider fragments as adjacent (default: 50)
    """
    if not fragments:
        return

    # Sort by baseline, then left position
    sorted_frags = sorted(fragments, key=lambda f: (f["baseline"], f["left"]))

    # Group fragments by baseline
    baseline_groups = []
    current_group = []
    current_baseline = None

    for f in sorted_frags:
        b = f["baseline"]
        if current_baseline is None:
            current_group = [f]
            current_baseline = b
        elif abs(b - current_baseline) <= baseline_tol:
            current_group.append(f)
        else:
            if current_group:
                baseline_groups.append(current_group)
            current_group = [f]
            current_baseline = b

    if current_group:
        baseline_groups.append(current_group)

    # Process each baseline group
    for group in baseline_groups:
        if len(group) <= 1:
            continue

        # Sort group by left position to process left-to-right
        group.sort(key=lambda f: f["left"])

        # Only propagate ColID 0 to horizontally ADJACENT fragments
        # Don't blindly set all fragments to ColID 0 if they're in different column territories
        for i in range(len(group) - 1):
            curr_frag = group[i]
            next_frag = group[i + 1]

            # Calculate horizontal gap between current and next fragment
            curr_right = curr_frag["left"] + curr_frag["width"]
            next_left = next_frag["left"]
            gap = next_left - curr_right

            # Only propagate ColID 0 if fragments are horizontally adjacent
            if gap <= max_word_gap:
                # If current is ColID 0 and next is not, propagate ColID 0 to next
                if curr_frag["col_id"] == 0 and next_frag["col_id"] != 0:
                    next_frag["col_id"] = 0
                # If next is ColID 0 and current is not, propagate ColID 0 backward
                elif next_frag["col_id"] == 0 and curr_frag["col_id"] != 0:
                    curr_frag["col_id"] = 0


def normalize_col_id_across_consecutive_lines(fragments, baseline_tol, left_tol=10):
    """
    Normalize ColID for lines that are clearly part of the same paragraph.

    When lines have the same left position (within left_tol), they should
    have consistent ColID values. This fixes cases where width-based thresholds cause
    inconsistent ColID assignment for paragraph lines with varying widths.

    This enhanced version tracks the last non-zero ColID seen for each left-position
    territory, allowing it to handle intervening elements from different columns.

    The logic:
    1. Group fragments by baseline into lines
    2. Track the last non-zero ColID seen for each left-position territory
    3. When a line has ColID=0, check if its left position matches a known territory
    4. If so, propagate that territory's ColID

    Args:
        fragments: List of fragments with baseline, left, and col_id
        baseline_tol: Tolerance for grouping into baselines
        left_tol: Tolerance for considering left positions as "same" (default: 10 pixels)
    """
    if not fragments:
        return

    # Sort by baseline, then left
    sorted_frags = sorted(fragments, key=lambda f: (f["baseline"], f["left"]))

    # Group fragments by baseline
    baseline_groups = []
    current_group = []
    current_baseline = None

    for f in sorted_frags:
        b = f["baseline"]
        if current_baseline is None:
            current_group = [f]
            current_baseline = b
        elif abs(b - current_baseline) <= baseline_tol:
            current_group.append(f)
            # Update to average baseline for better grouping
            current_baseline = sum(frag["baseline"] for frag in current_group) / len(current_group)
        else:
            if current_group:
                baseline_groups.append(current_group)
            current_group = [f]
            current_baseline = b

    if current_group:
        baseline_groups.append(current_group)

    # Track the last non-zero ColID seen for each left-position territory
    # Key: rounded left position (to left_tol precision), Value: last non-zero ColID
    left_territory_col_ids = {}

    def get_territory_key(left_pos):
        """Round left position to create territory buckets."""
        return round(left_pos / left_tol) * left_tol

    # First pass: build the territory map from lines with non-zero ColID
    for group in baseline_groups:
        group_left = min(f["left"] for f in group)
        territory_key = get_territory_key(group_left)

        # Get non-zero ColID from this group
        non_zero_col_ids = [f["col_id"] for f in group if f["col_id"] > 0]
        if non_zero_col_ids:
            # Use the most common non-zero ColID
            col_id = max(set(non_zero_col_ids), key=non_zero_col_ids.count)
            left_territory_col_ids[territory_key] = col_id

    # Second pass: propagate ColID to lines with ColID=0 based on territory
    for group in baseline_groups:
        group_left = min(f["left"] for f in group)
        territory_key = get_territory_key(group_left)

        # Check if this group has ColID=0 fragments
        col0_frags = [f for f in group if f["col_id"] == 0]
        if not col0_frags:
            continue

        # Look for a matching territory ColID
        # Check exact key and nearby keys (within left_tol)
        target_col_id = None
        for offset in [0, -left_tol, left_tol]:
            check_key = territory_key + offset
            if check_key in left_territory_col_ids:
                target_col_id = left_territory_col_ids[check_key]
                break

        # If we found a matching territory, propagate the ColID
        if target_col_id is not None:
            for f in col0_frags:
                f["col_id"] = target_col_id

    # Third pass: handle adjacent lines (original logic as fallback)
    for i in range(len(baseline_groups) - 1):
        curr_group = baseline_groups[i]
        next_group = baseline_groups[i + 1]

        # Get leftmost position for each group
        curr_left = min(f["left"] for f in curr_group)
        next_left = min(f["left"] for f in next_group)

        # Skip if left positions are significantly different (different columns)
        if abs(curr_left - next_left) > left_tol:
            continue

        # Get ColID values for each group
        curr_col_ids = {f["col_id"] for f in curr_group}
        next_col_ids = {f["col_id"] for f in next_group}

        # If groups have different ColID but same left position, unify them
        # Preference: If first group has ColID > 0, propagate it to ColID=0 groups
        # This handles the case where paragraph text is incorrectly split due to width threshold
        if curr_col_ids != next_col_ids:
            # Get the dominant ColID from current group (prefer non-zero)
            curr_col_id = next(iter(c for c in curr_col_ids if c > 0), 0)
            next_col_id = next(iter(c for c in next_col_ids if c > 0), 0)

            # If current line has a specific column and next line is "full-width",
            # the next line probably should also be in that column
            if curr_col_id > 0 and next_col_id == 0:
                for f in next_group:
                    if f["col_id"] == 0:
                        f["col_id"] = curr_col_id

            # Also handle the reverse case: if next line has a specific column
            # and current line is "full-width", update current line
            elif next_col_id > 0 and curr_col_id == 0:
                for f in curr_group:
                    if f["col_id"] == 0:
                        f["col_id"] = next_col_id


def fix_isolated_centered_fragments(fragments, page_width, col_starts=None):
    """
    Fix center-aligned fragments that got assigned wrong ColID based on position.

    Small centered fragments (like centered titles, subtitles) might get ColID 2
    based on their left position falling in column 2's territory. This function
    detects such fragments and reassigns them to match their surrounding context.

    NOTE: This function only applies to SINGLE-COLUMN pages. On multi-column pages,
    if a fragment has ColID=2, it's because it's actually in column 2.

    Detection criteria for centered content:
    - Fragment center is VERY close to page center (within 10% of page width)
    - Fragment is relatively narrow (less than 30% of page width)
    - Fragment's left position does NOT align with typical column starts
    - Fragment's ColID is different from the previous fragment on the page

    If detected, inherit ColID from the previous fragment (or use ColID 0/1 as fallback).
    """
    if not fragments or len(fragments) < 2:
        return

    # Skip for multi-column pages - ColID assignments are intentional
    if col_starts and len(col_starts) > 1:
        return

    page_center = page_width / 2.0
    # Use tighter tolerance to avoid catching column 2 content
    center_tolerance = page_width * 0.10  # 10% tolerance (was 20%)
    narrow_threshold = page_width * 0.30   # Consider narrow if < 30% (was 40%)

    # Typical column start positions (as ratio of page width)
    # Column 1 typically starts around 10-15% from left
    # Column 2 typically starts around 50-55% from left
    col1_zone = (page_width * 0.10, page_width * 0.20)  # 10-20% of page
    col2_zone = (page_width * 0.45, page_width * 0.60)  # 45-60% of page

    # Sort fragments by baseline (top to bottom), then left
    sorted_frags = sorted(fragments, key=lambda f: (f.get("norm_baseline", f["baseline"]), f["left"]))

    # Track previous fragment for context
    prev_frag = None

    for f in sorted_frags:
        if prev_frag is None:
            prev_frag = f
            continue

        # Check if this fragment looks like isolated centered content
        fragment_center = f["left"] + f["width"] / 2.0
        is_centered = abs(fragment_center - page_center) <= center_tolerance
        is_narrow = f["width"] < narrow_threshold
        has_different_col = f["col_id"] != prev_frag["col_id"]

        # Check if fragment's left position is in a typical column zone
        # If so, it's likely real column content, not centered text
        in_column_zone = (col1_zone[0] <= f["left"] <= col1_zone[1] or
                         col2_zone[0] <= f["left"] <= col2_zone[1])

        # Skip if fragment is already ColID 0 or 1 (those are usually correct)
        if f["col_id"] <= 1:
            prev_frag = f
            continue

        # Skip if fragment is in a typical column zone - it's real column content
        if in_column_zone:
            prev_frag = f
            continue

        # If fragment is clearly centered (tight tolerance), inherit ColID from context
        if is_centered and is_narrow:
            # Clearly centered content - reassign based on context
            if prev_frag["col_id"] in (0, 1):
                f["col_id"] = prev_frag["col_id"]
            else:
                # If previous is also a higher column, default to ColID 0 (full-width)
                # since centered content should generally be treated as full-width
                f["col_id"] = 0
            prev_frag = f
            continue

        # If fragment has different ColID from context and is not clearly a column fragment,
        # it might be misclassified - inherit from previous
        if has_different_col:
            if prev_frag["col_id"] in (0, 1):
                f["col_id"] = prev_frag["col_id"]
            else:
                f["col_id"] = 0

        prev_frag = f


def reclassify_footnote_rows_as_fullwidth(rows, page_width, page_height):
    """
    Detect and reclassify footnote rows that span multiple columns.

    Footnotes are often broken into multiple text fragments on the same baseline,
    where each fragment individually gets assigned to different columns, but
    collectively they form a full-width line at the bottom of the page.

    This function handles two cases:
    1. Rows with fragments from multiple columns that collectively span full-width
    2. Individual wide fragments (after merging) that should be reclassified as full-width

    This function detects such rows and reclassifies all fragments as col_id=0.
    """
    if not rows:
        return

    # Focus on bottom 25% of page where footnotes typically appear
    footnote_threshold = page_height * 0.75

    # Width threshold for considering a fragment as full-width
    width_threshold = page_width * 0.60  # 60% of page width

    for row in rows:
        if not row:
            continue

        # Check if row is in the footnote area (bottom 25%)
        row_top = min(f["top"] for f in row)
        if row_top < footnote_threshold:
            continue

        # Calculate collective span of all fragments in this row
        row_left = min(f["left"] for f in row)
        row_right = max(f["left"] + f["width"] for f in row)
        row_span = row_right - row_left

        # Get unique column IDs (excluding 0 which is already full-width)
        col_ids = {f["col_id"] for f in row if f["col_id"] != 0}

        # Case 1: Row has fragments from multiple columns spanning >75% page width
        # Case 2: Row collectively spans >75% page width (even if same column after merging)
        if row_span >= page_width * 0.75:
            # Reclassify all fragments in this row as full-width
            for f in row:
                f["col_id"] = 0
        # Case 3: Single wide fragment (after merging) that should be full-width
        elif len(row) == 1 and row[0]["width"] >= width_threshold and row[0]["col_id"] != 0:
            row[0]["col_id"] = 0


# -------------------------------------------------------------
# Main PDF → Excel conversion
# -------------------------------------------------------------
def pdf_to_excel_with_columns(
    pdf_path,
    pdftohtml_xml_path=None,
    excel_output_path=None,
    exclusion_rects=None,
    pymupdf_page_dims=None,
):
    """
    Convert PDF to Excel with column detection and reading order assignment.

    Args:
        pdf_path: Path to the input PDF file
        pdftohtml_xml_path: Optional path to existing pdftohtml XML output
        excel_output_path: Optional path for Excel output
        exclusion_rects: Optional dict mapping page_no (1-indexed) to list of
                        (x0, y0, x1, y1) bboxes in PyMuPDF coordinates.
                        Text inside these regions will be filtered BEFORE
                        column detection to prevent table text from distorting
                        the reading order.
        pymupdf_page_dims: Optional dict mapping page_no (1-indexed) to
                          (page_width, page_height) in PyMuPDF coordinates.
                          Required for scaling exclusion_rects to pdftohtml space.
    """
    # 1) Run pdftohtml -xml if needed
    if pdftohtml_xml_path is None:
        base, _ = os.path.splitext(pdf_path)
        pdftohtml_xml_path = base + "_pdftohtml.xml"

    if not os.path.exists(pdftohtml_xml_path):
        pdftohtml_xml_path = run_pdftohtml_xml(pdf_path, pdftohtml_xml_path)
    else:
        print(f"Using existing pdftohtml XML: {pdftohtml_xml_path}")

    # 1a) Backup original XML before we start processing
    base_xml, _ = os.path.splitext(pdftohtml_xml_path)
    backup_xml_path = base_xml + "_original.xml"
    if not os.path.exists(backup_xml_path):
        try:
            with open(pdftohtml_xml_path, "rb") as src, open(
                backup_xml_path, "wb"
            ) as dst:
                dst.write(src.read())
            print(f"Backed up original XML to: {backup_xml_path}")
        except Exception as e:
            print(f"Warning: could not back up XML: {e}")

    # 2) Parse XML
    tree = ET.parse(pdftohtml_xml_path)
    root = tree.getroot()

    # Prepare Excel workbook
    wb = Workbook()
    ws_ro = wb.active
    ws_ro.title = "ReadingOrder"

    ws_lines = wb.create_sheet("Lines")
    ws_img = wb.create_sheet("Images")
    ws_debug = wb.create_sheet("Debug")

    # Headers
    ws_ro.append(
        [
            "Page",
            "StreamIndex",
            "ReadingOrderBlock",
            "ColID",
            "RowIndex",
            "Left",
            "Top",
            "Width",
            "Height",
            "Baseline",
            "NormBaseline",
            "Text",
        ]
    )

    ws_lines.append(
        [
            "Page",
            "RowIndex",
            "Baseline",
            "Col0_Text",
            "Col1_Text",
            "Col2_Text",
            "Col3_Text",
            "Col4_Text",
        ]
    )

    ws_img.append(["Page", "ImageIndex", "Left", "Top", "Width", "Height", "Label"])

    ws_debug.append(
        [
            "Page",
            "StreamIndex",
            "ColID",
            "RowIndex",
            "Baseline",
            "Left",
            "Top",
            "Width",
            "Height",
            "NormText",
        ]
    )

    seen_footer_texts = set()

    # Pre-pass: Scan pages to identify repeated header/footer text (chapter titles, page numbers)
    # Text that appears in same position on multiple pages is likely header/footer noise
    header_footer_candidates = {}  # {(norm_position, norm_text): count}
    page_elements_prescan = list(root.findall(".//page"))

    print(f"Pre-scanning {len(page_elements_prescan)} pages for header/footer patterns...")
    for page_elem in page_elements_prescan:
        page_height = float(page_elem.get("height", "0") or 0.0)
        page_width = float(page_elem.get("width", "0") or 0.0)

        if page_height <= 0:
            continue

        for t in page_elem.findall("text"):
            txt_raw = "".join(t.itertext())
            norm_txt = " ".join(txt_raw.split()).lower()
            
            # FIX 1: Add minimum text length requirement (5 chars)
            # Skip very short text (bullets, single chars) and very long text
            if not norm_txt or len(norm_txt) < 5 or len(norm_txt) > 100:
                continue
            
            # FIX 2: Exclude figure/table labels - they're legitimate content, not headers/footers
            if re.match(r'^(figure|table|fig\.?)\s+\d+', norm_txt, re.IGNORECASE):
                continue

            top = float(t.get("top", "0") or 0.0)
            left = float(t.get("left", "0") or 0.0)

            # Check if in header zone (top 12%) or footer zone (bottom 12%)
            norm_top = round(top / page_height, 2) if page_height > 0 else 0
            norm_left = round(left / page_width, 2) if page_width > 0 else 0

            is_header_zone = norm_top < 0.12
            is_footer_zone = norm_top > 0.88

            if is_header_zone or is_footer_zone:
                # Create a position key (rounded position + text)
                pos_key = (round(norm_top, 1), round(norm_left, 1), norm_txt)
                header_footer_candidates[pos_key] = header_footer_candidates.get(pos_key, 0) + 1

    # FIX 3: Increase minimum occurrence threshold for large documents
    # For large documents (1000+ pages), require higher repetition (at least 10 occurrences or 1% of pages)
    # For small documents, keep threshold at 3
    if len(page_elements_prescan) >= 500:
        min_occurrences = max(10, len(page_elements_prescan) // 100)  # At least 10, or 1% of pages
    else:
        min_occurrences = max(3, len(page_elements_prescan) // 10)  # At least 3, or 10% of pages
    
    print(f"  Using minimum occurrence threshold: {min_occurrences} (for {len(page_elements_prescan)} pages)")
    
    for (norm_top, norm_left, norm_txt), count in header_footer_candidates.items():
        if count >= min_occurrences:
            seen_footer_texts.add(norm_txt)
            print(f"  Header/footer pattern detected ({count}x): '{norm_txt[:50]}...' at position ({norm_top}, {norm_left})")

    if seen_footer_texts:
        print(f"  Total header/footer patterns to filter: {len(seen_footer_texts)}")
    else:
        print(f"  No repeated header/footer patterns detected")

    # Store all page data for return
    all_pages_data = {}
    
    # NEW: Track page order for cross-page merging
    page_order = []

    # Iterate over pages
    page_elements = list(root.findall(".//page"))
    total_pages = len(page_elements)
    print(f"Processing {total_pages} pages...")
    
    for page_idx, page_elem in enumerate(page_elements, 1):
        page_number = int(page_elem.get("number", "0") or 0)
        
        # Progress indicator every 50 pages
        if page_idx % 50 == 0 or page_idx == 1:
            print(f"  Processing page {page_idx}/{total_pages} (page number: {page_number})")
        page_width = float(page_elem.get("width", "0") or 0.0)
        page_height = float(page_elem.get("height", "0") or 0.0)

        # Images → simple placeholders
        img_idx = 1
        for img in page_elem.findall("image"):
            l = float(img.get("left", "0") or 0.0)
            t = float(img.get("top", "0") or 0.0)
            w = float(img.get("width", "0") or 0.0)
            h = float(img.get("height", "0") or 0.0)
            ws_img.append([page_number, img_idx, l, t, w, h, "IMAGE_PLACEHOLDER"])
            img_idx += 1

        # Collect text fragments
        fragments = []
        page_number_fragments = []  # FIX 4: Separate list for page numbers (for page ID extraction)
        stream_index = 1

        for t in page_elem.findall("text"):
            # capture inner <b>, <i> etc.
            txt_raw = "".join(t.itertext())   # Get plain text for display/filtering
            txt = txt_raw                     # no strip
            if not txt:
                continue

            # Preserve inner XML structure for formatting (stores as XML string)
            inner_xml = ET.tostring(t, encoding="unicode", method="xml")
            # Remove the outer <text...> wrapper, keeping only inner content
            # Extract content between opening and closing <text> tags
            inner_content = inner_xml
            if inner_xml.startswith("<text"):
                start = inner_xml.find(">") + 1
                end = inner_xml.rfind("</text>")
                if start > 0 and end > start:
                    inner_content = inner_xml[start:end]
                elif inner_xml.endswith("/>"):
                    # Self-closing tag, no content
                    inner_content = txt_raw
                else:
                    inner_content = txt_raw
            else:
                inner_content = txt_raw

            left = float(t.get("left", "0") or 0.0)
            top = float(t.get("top", "0") or 0.0)
            width = float(t.get("width", "0") or 0.0)
            height = float(t.get("height", "0") or 0.0)
            baseline = top + height

            # Try a few common attribute names for rotation; fall back to 0 if missing
            rot_raw = t.get("rotation") or t.get("rotate") or t.get("rot") or "0"
            try:
                rotation_deg = int(float(rot_raw))
            except ValueError:
                rotation_deg = 0

            # Check for vertical spine text to skip
            # Filter out vertical spine text at page border (INTRODUCTION, etc.)
            if is_vertical_spine_text(txt, left, width, height,
                              page_width, page_height, rotation_deg):
                continue

            norm_txt = " ".join(txt.split()).lower()
            
            # FIX 4: Check if this is a standalone page number BEFORE filtering
            # Preserve page numbers for page ID extraction even if they'd be filtered from content
            is_page_number = False
            if page_height > 0:
                norm_top = top / page_height
                is_header_zone = norm_top < 0.12
                is_footer_zone = norm_top > 0.85
                
                if is_header_zone or is_footer_zone:
                    text_stripped = norm_txt.strip()
                    # Check for arabic numbers (1-9999) or roman numerals
                    if re.match(r'^\d{1,4}$', text_stripped) or re.match(r'^[ivxlcdm]+$', text_stripped, re.IGNORECASE):
                        is_page_number = True
                        # Store in separate list for page ID extraction
                        page_number_fragments.append({
                            "text": txt,
                            "norm_text": norm_txt,
                            "left": left,
                            "top": top,
                            "width": width,
                            "height": height,
                        })
            
            # Apply normal filtering (page numbers will be filtered from main content but preserved above)
            if should_skip_fragment(norm_txt, top, height, page_height, seen_footer_texts):
                continue

            fragments.append({
                "stream_index": stream_index,
                "text": txt,
                "inner_xml": inner_content,  # Preserve formatting tags
                "norm_text": norm_txt,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "baseline": baseline,
                "col_id": None,
                "row_index": None,
                "reading_order_block": None,
            })
            stream_index += 1

        if not fragments:
            continue

        # ===== Filter out text inside exclusion rects (tables) =====
        # This MUST happen BEFORE column detection to prevent table text
        # from distorting the reading order and column assignment.
        # IMPORTANT: exclusion_rects are in PyMuPDF coordinates, but fragments
        # are in pdftohtml coordinates. We must scale to compare correctly.
        if exclusion_rects and page_number in exclusion_rects:
            page_exclusions = exclusion_rects[page_number]
            if page_exclusions:
                # Calculate scale factors: PyMuPDF coords -> pdftohtml coords
                # pdftohtml_coord = pymupdf_coord * (html_size / pymupdf_size)
                scale_x = 1.0
                scale_y = 1.0
                if pymupdf_page_dims and page_number in pymupdf_page_dims:
                    pymupdf_w, pymupdf_h = pymupdf_page_dims[page_number]
                    if pymupdf_w > 0 and page_width > 0:
                        scale_x = page_width / pymupdf_w
                    if pymupdf_h > 0 and page_height > 0:
                        scale_y = page_height / pymupdf_h
                    if abs(scale_x - 1.0) > 0.01 or abs(scale_y - 1.0) > 0.01:
                        print(f"  Page {page_number}: Scale factors for coordinate conversion: "
                              f"x={scale_x:.3f}, y={scale_y:.3f}")

                # Scale exclusion rects from PyMuPDF to pdftohtml space
                scaled_exclusions = []
                for (ex_x0, ex_y0, ex_x1, ex_y1) in page_exclusions:
                    scaled_exclusions.append((
                        ex_x0 * scale_x,
                        ex_y0 * scale_y,
                        ex_x1 * scale_x,
                        ex_y1 * scale_y,
                    ))

                original_count = len(fragments)
                filtered_fragments = []
                for frag in fragments:
                    # Calculate fragment bbox (already in pdftohtml coords)
                    f_x0 = frag["left"]
                    f_y0 = frag["top"]
                    f_x1 = f_x0 + frag["width"]
                    f_y1 = f_y0 + frag["height"]

                    # Check if fragment is inside any exclusion rect (now in same coord space)
                    inside_exclusion = False
                    for (ex_x0, ex_y0, ex_x1, ex_y1) in scaled_exclusions:
                        # Check if fragment center is inside exclusion rect
                        f_cx = (f_x0 + f_x1) / 2
                        f_cy = (f_y0 + f_y1) / 2
                        if ex_x0 <= f_cx <= ex_x1 and ex_y0 <= f_cy <= ex_y1:
                            inside_exclusion = True
                            break

                    if not inside_exclusion:
                        filtered_fragments.append(frag)

                excluded_count = original_count - len(filtered_fragments)
                if excluded_count > 0:
                    print(f"  Page {page_number}: Excluded {excluded_count} fragments inside table regions")
                fragments = filtered_fragments

                if not fragments:
                    continue

        # Debug: Log if page numbers were found
        if page_number_fragments:
            print(f"  Page {page_number}: Found {len(page_number_fragments)} page number(s) for ID extraction")

        # Sort by baseline & left for line grouping
        fragments.sort(key=lambda f: (f["baseline"], f["left"]))
        
        # ===== Phase 1: Detect superscripts/subscripts =====
        # Detect and mark scripts BEFORE grouping into rows.
        # Uses TOP position (not baseline) to find scripts adjacent to larger text.
        # Very strict criteria (w<15, h<12) to avoid false positives (drop caps, etc.)
        script_count = detect_and_mark_scripts(fragments)
        if script_count > 0:
            print(f"  Page {page_number}: Detected {script_count} superscript(s)/subscript(s)")

        # Warn if page has many fragments (potential performance issue)
        if len(fragments) > 1000:
            print(f"  Page {page_number}: {len(fragments)} fragments (large page, may take longer)")

        # FIRST: Group fragments into lines and compute norm_baseline
        # This must happen BEFORE column detection so we can use accurate baselines
        baselines_pre = [f["baseline"] for f in fragments]
        baseline_tol_pre = compute_baseline_tolerance(baselines_pre) if baselines_pre else 2.0
        pre_rows = group_fragments_into_lines(fragments, baseline_tol_pre)
        assign_normalized_baselines(pre_rows)

        # Column detection using norm_baseline for accurate line grouping
        col_starts = detect_column_starts(fragments, page_width, max_cols=4)
        assign_column_ids(fragments, page_width, col_starts)

        # Reassign misclassified ColID 0 fragments to correct columns
        reassign_misclassified_col0_fragments(fragments, page_width, col_starts)

        # Maintain ColID 0 for all fragments on the same baseline
        # (fixes issue where last small fragment gets assigned to wrong column)
        baselines_for_col0 = [f["baseline"] for f in fragments]
        baseline_tol_for_col0 = compute_baseline_tolerance(baselines_for_col0) if baselines_for_col0 else 2.0
        maintain_col0_within_baseline(fragments, baseline_tol_for_col0)

        # Normalize ColID across consecutive lines with same left position
        # (fixes issue where paragraph lines get different ColID due to width threshold)
        normalize_col_id_across_consecutive_lines(fragments, baseline_tol_for_col0)

        # Fix isolated centered fragments that got wrong ColID based on position
        # (e.g., centered titles getting ColID 2 instead of inheriting from context)
        # Only applies to single-column pages - multi-column pages have intentional ColID assignments
        fix_isolated_centered_fragments(fragments, page_width, col_starts)

        # (1) First pass: group into rows and merge inline fragments within each row
        baselines = [f["baseline"] for f in fragments]
        baseline_tol = compute_baseline_tolerance(baselines)
        raw_rows = group_fragments_into_lines(fragments, baseline_tol)
        
        # ===== Phase 3: Merge scripts across rows =====
        # Merge superscripts/subscripts with their parents even if in different rows.
        # This fixes cases like "10^7" where "7" is in a different row due to baseline difference.
        raw_rows = merge_scripts_across_rows(raw_rows, fragments)

        merged_fragments = []
        for row in raw_rows:
            merged_fragments.extend(merge_inline_fragments_in_row(row))

        fragments = merged_fragments
        if not fragments:
            continue

        # Re-sort after merging
        fragments.sort(key=lambda f: (f["baseline"], f["left"]))

        # (2) Now group again into rows with merged fragments
        baselines = [f["baseline"] for f in fragments]
        baseline_tol = compute_baseline_tolerance(baselines)
        rows = group_fragments_into_lines(fragments, baseline_tol)

        # (2) Assign normalized baselines to each fragment (for baseline-driven block assignment)
        assign_normalized_baselines(rows)

        # (2a) Reclassify footnote rows that span multiple columns as full-width
        # DISABLED: The ColID logic is now robust enough without special footer zone handling
        # This was incorrectly reclassifying normal column text in bottom 25% of pages as ColID=0
        # reclassify_footnote_rows_as_fullwidth(rows, page_width, page_height)

        # (2b) Group ColID 0 fragments by vertical gap - only for wide fragments
        # Calculate typical line height for gap detection
        line_heights = [f["height"] for f in fragments if f["height"] > 0]
        typical_line_height = sorted(line_heights)[len(line_heights) // 2] if line_heights else 12.0
        group_col0_by_vertical_gap(fragments, typical_line_height, page_width=page_width, page_height=page_height)

        # (3) assign row indices
        row_idx = 1
        for row in rows:
            for f in row:
                f["row_index"] = row_idx
            row_idx += 1

        # (4) Assign reading-order blocks (using norm_baseline for ordering)
        assign_reading_order_blocks(fragments, rows)

        # Store page data for return (copy fragments to preserve all info)
        all_pages_data[page_number] = {
            "page_width": page_width,
            "page_height": page_height,
            "fragments": [dict(f) for f in fragments],  # deep copy
            "page_number_fragments": page_number_fragments,  # FIX 4: Preserve page numbers for page ID extraction
        }
        
        # Track page order for cross-page merging
        page_order.append(page_number)

    # ---------------------------------------------------------
    # Cross-page paragraph merging
    # ---------------------------------------------------------
    print(f"\nCompleted processing all {total_pages} pages")
    
    # Merge paragraphs that continue across pages
    all_pages_data = merge_paragraphs_across_pages(all_pages_data, page_order)
    
    # ---------------------------------------------------------
    # Write Excel output (after cross-page merging)
    # ---------------------------------------------------------
    print("Writing Excel output...")
    
    for page_num in page_order:
        if page_num not in all_pages_data:
            continue
        
        page_data = all_pages_data[page_num]
        fragments = page_data["fragments"]
        
        # Group fragments into rows for Lines sheet
        fragments_sorted = sorted(fragments, key=lambda f: (f["baseline"], f["left"]))
        baselines = [f["baseline"] for f in fragments_sorted]
        baseline_tol = compute_baseline_tolerance(baselines) if baselines else 2.0
        rows = group_fragments_into_lines(fragments_sorted, baseline_tol)
        
        # Write ReadingOrder and Debug sheets
        for f in fragments:
            # Skip fragments that were merged to previous page
            if f.get("merged_to_prev_page"):
                continue
            
            ws_ro.append(
                [
                    page_num,
                    f["stream_index"],
                    f["reading_order_block"],
                    f["col_id"],
                    f["row_index"],
                    f["left"],
                    f["top"],
                    f["width"],
                    f["height"],
                    f["baseline"],
                    f.get("norm_baseline", f["baseline"]),
                    f["text"],
                ]
            )
            ws_debug.append(
                [
                    page_num,
                    f["stream_index"],
                    f["col_id"],
                    f["row_index"],
                    f["baseline"],
                    f["left"],
                    f["top"],
                    f["width"],
                    f["height"],
                    f["norm_text"],
                ]
            )
        
        # Write Lines sheet (grouped by row & col)
        for row in rows:
            if not row:
                continue
            
            # Skip rows where all fragments are merged to prev page
            active_frags = [f for f in row if not f.get("merged_to_prev_page")]
            if not active_frags:
                continue

            b_row = statistics.mean(f["baseline"] for f in active_frags)
            row_index = active_frags[0]["row_index"]

            # col_id -> list of frags in that row
            by_col = {}
            for f in active_frags:
                c = f["col_id"]
                by_col.setdefault(c, []).append(f)

            col_ids = sorted(by_col.keys())

            # up to 5 buckets: 0,1,2,3,4
            texts = ["", "", "", "", ""]
            for c in col_ids:
                seg_frags = sorted(by_col[c], key=lambda f: f["left"])
                seg_text = " ".join(f["text"] for f in seg_frags)
                idx_c = c if 0 <= c <= 4 else 4
                # Append with separator if multiple segments for same col
                if texts[idx_c]:
                    texts[idx_c] += " | " + seg_text
                else:
                    texts[idx_c] = seg_text

            ws_lines.append(
                [
                    page_num,
                    row_index,
                    b_row,
                    texts[0],
                    texts[1],
                    texts[2],
                    texts[3],
                    texts[4],
                ]
            )
    
    # ---------------------------------------------------------
    # Save Excel
    # ---------------------------------------------------------
    print("Saving Excel file...")
    
    if excel_output_path is None:
        base, _ = os.path.splitext(pdf_path)
        excel_output_path = base + "_columns.xlsx"

    wb.save(excel_output_path)
    print(f"✓ Excel saved to: {excel_output_path}")

    # Return structured data for unified XML generation
    return {
        "excel_path": excel_output_path,
        "pdftohtml_xml_path": pdftohtml_xml_path,
        "pages": all_pages_data,
    }


# -------------------------------------------------------------
# CLI
# -------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert PDF (via pdftohtml -xml) into Excel with column-aware reading order."
    )
    parser.add_argument("pdf_path", help="Path to input PDF")
    parser.add_argument("--xml", dest="pdftohtml_xml_path", help="Output XML path")
    parser.add_argument("--excel", dest="excel_output_path", help="Output Excel path")

    args = parser.parse_args()

    pdf_to_excel_with_columns(
        pdf_path=args.pdf_path,
        pdftohtml_xml_path=args.pdftohtml_xml_path,
        excel_output_path=args.excel_output_path,
    )
