#!/usr/bin/env python3
"""
MongoDB Integration for Conversion Dashboard

This module provides MongoDB storage for conversion tracking data,
enabling persistent dashboard and reporting capabilities for the UI.

Features:
- Automatic data push on job completion
- Dashboard statistics queries
- Historical reporting
- Aggregation pipelines for analytics

Configuration:
    Set these environment variables:
    - MONGODB_URI: MongoDB connection string (default: mongodb://localhost:27017)
    - MONGODB_DATABASE: Database name (default: pdftoxml)
    - MONGODB_COLLECTION: Collection name (default: conversion_dashboard)

Usage:
    from mongodb_store import MongoDBStore, get_mongodb_store

    # Get singleton instance
    store = get_mongodb_store()

    # Push conversion data
    store.push_conversion(conversion_data)

    # Get dashboard stats
    stats = store.get_dashboard_stats()
"""

from __future__ import annotations

import logging
import os
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# Check for pymongo availability
try:
    from pymongo import MongoClient, DESCENDING
    from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError
    from bson import ObjectId
    MONGODB_AVAILABLE = True
except ImportError:
    MONGODB_AVAILABLE = False
    MongoClient = None
    logger.warning("pymongo not installed. MongoDB features will be disabled.")


# ============================================================================
# CONFIGURATION
# ============================================================================

@dataclass
class MongoDBConfig:
    """MongoDB connection configuration."""
    uri: str = "mongodb://localhost:27017"
    database: str = "pdftoxml"
    collection: str = "conversion_dashboard"
    timeout_ms: int = 5000
    retry_writes: bool = True

    @classmethod
    def from_env(cls) -> "MongoDBConfig":
        """Load configuration from environment variables."""
        return cls(
            uri=os.environ.get("MONGODB_URI", "mongodb://localhost:27017"),
            database=os.environ.get("MONGODB_DATABASE", "pdftoxml"),
            collection=os.environ.get("MONGODB_COLLECTION", "conversion_dashboard"),
            timeout_ms=int(os.environ.get("MONGODB_TIMEOUT_MS", "5000")),
        )


# ============================================================================
# DATA MODELS
# ============================================================================

@dataclass
class ConversionDocument:
    """MongoDB document structure for a conversion record."""
    # Identifiers
    job_id: str
    filename: str
    isbn: Optional[str] = None
    title: Optional[str] = None
    publisher: Optional[str] = None
    authors: List[str] = None

    # Timing
    start_time: datetime = None
    end_time: datetime = None
    duration_seconds: Optional[float] = None

    # Status
    status: str = "in_progress"  # in_progress, success, failure, partial
    progress_percent: int = 0
    error_message: Optional[str] = None

    # Source info
    conversion_type: str = "PDF"  # PDF, ePub, DOCX, etc.
    template_type: str = "unknown"  # single_column, double_column, mixed

    # Statistics
    num_chapters: int = 0
    num_pages: int = 0
    num_vector_images: int = 0
    num_raster_images: int = 0
    num_total_images: int = 0
    num_tables: int = 0
    num_equations: int = 0

    # Output
    output_path: Optional[str] = None
    output_size_mb: Optional[float] = None
    docx_path: Optional[str] = None
    package_path: Optional[str] = None

    # Metadata
    created_at: datetime = None
    updated_at: datetime = None

    def __post_init__(self):
        if self.authors is None:
            self.authors = []
        if self.created_at is None:
            self.created_at = datetime.utcnow()
        if self.updated_at is None:
            self.updated_at = datetime.utcnow()
        if self.start_time is None:
            self.start_time = datetime.utcnow()
        self.num_total_images = self.num_vector_images + self.num_raster_images

    def to_dict(self) -> Dict[str, Any]:
        """Convert to MongoDB document format."""
        data = {
            "job_id": self.job_id,
            "filename": self.filename,
            "isbn": self.isbn,
            "title": self.title,
            "publisher": self.publisher,
            "authors": self.authors,
            "start_time": self.start_time,
            "end_time": self.end_time,
            "duration_seconds": self.duration_seconds,
            "status": self.status,
            "progress_percent": self.progress_percent,
            "error_message": self.error_message,
            "conversion_type": self.conversion_type,
            "template_type": self.template_type,
            "num_chapters": self.num_chapters,
            "num_pages": self.num_pages,
            "num_vector_images": self.num_vector_images,
            "num_raster_images": self.num_raster_images,
            "num_total_images": self.num_total_images,
            "num_tables": self.num_tables,
            "num_equations": self.num_equations,
            "output_path": self.output_path,
            "output_size_mb": self.output_size_mb,
            "docx_path": self.docx_path,
            "package_path": self.package_path,
            "created_at": self.created_at,
            "updated_at": self.updated_at,
        }
        return data

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "ConversionDocument":
        """Create from MongoDB document."""
        return cls(
            job_id=data.get("job_id", ""),
            filename=data.get("filename", ""),
            isbn=data.get("isbn"),
            title=data.get("title"),
            publisher=data.get("publisher"),
            authors=data.get("authors", []),
            start_time=data.get("start_time"),
            end_time=data.get("end_time"),
            duration_seconds=data.get("duration_seconds"),
            status=data.get("status", "in_progress"),
            progress_percent=data.get("progress_percent", 0),
            error_message=data.get("error_message"),
            conversion_type=data.get("conversion_type", "PDF"),
            template_type=data.get("template_type", "unknown"),
            num_chapters=data.get("num_chapters", 0),
            num_pages=data.get("num_pages", 0),
            num_vector_images=data.get("num_vector_images", 0),
            num_raster_images=data.get("num_raster_images", 0),
            num_total_images=data.get("num_total_images", 0),
            num_tables=data.get("num_tables", 0),
            num_equations=data.get("num_equations", 0),
            output_path=data.get("output_path"),
            output_size_mb=data.get("output_size_mb"),
            docx_path=data.get("docx_path"),
            package_path=data.get("package_path"),
            created_at=data.get("created_at"),
            updated_at=data.get("updated_at"),
        )


# ============================================================================
# MONGODB STORE
# ============================================================================

class MongoDBStore:
    """
    MongoDB storage for conversion tracking data.

    Provides methods to:
    - Push conversion data on job completion
    - Query dashboard statistics
    - Generate reports and analytics
    """

    def __init__(self, config: Optional[MongoDBConfig] = None):
        """
        Initialize MongoDB store.

        Args:
            config: MongoDB configuration. If None, loads from environment.
        """
        self.config = config or MongoDBConfig.from_env()
        self._client: Optional[MongoClient] = None
        self._db = None
        self._collection = None
        self._connected = False

    def connect(self) -> bool:
        """
        Establish connection to MongoDB.

        Returns:
            True if connected successfully, False otherwise.
        """
        if not MONGODB_AVAILABLE:
            logger.error("pymongo is not installed. Run: pip install pymongo")
            return False

        try:
            self._client = MongoClient(
                self.config.uri,
                serverSelectionTimeoutMS=self.config.timeout_ms,
                retryWrites=self.config.retry_writes,
            )
            # Test connection
            self._client.admin.command('ping')

            self._db = self._client[self.config.database]
            self._collection = self._db[self.config.collection]

            # Create indexes for efficient queries
            self._create_indexes()

            self._connected = True
            logger.info(f"Connected to MongoDB: {self.config.database}.{self.config.collection}")
            return True

        except (ConnectionFailure, ServerSelectionTimeoutError) as e:
            logger.error(f"Failed to connect to MongoDB: {e}")
            self._connected = False
            return False
        except Exception as e:
            logger.error(f"Unexpected error connecting to MongoDB: {e}")
            self._connected = False
            return False

    def _create_indexes(self):
        """Create indexes for efficient queries."""
        if self._collection is not None:
            self._collection.create_index("job_id", unique=True)
            self._collection.create_index("isbn", unique=True, sparse=True)  # ISBN index for fast lookups
            self._collection.create_index("status")
            self._collection.create_index("created_at")
            self._collection.create_index([("created_at", DESCENDING)])
            self._collection.create_index("filename")
            logger.debug("MongoDB indexes created")

    def disconnect(self):
        """Close MongoDB connection."""
        if self._client:
            self._client.close()
            self._connected = False
            logger.info("Disconnected from MongoDB")

    @property
    def is_connected(self) -> bool:
        """Check if connected to MongoDB."""
        return self._connected and self._client is not None

    def ensure_connected(self) -> bool:
        """Ensure connection is established, reconnecting if needed."""
        if not self.is_connected:
            return self.connect()
        return True

    # ========================================================================
    # DATA OPERATIONS
    # ========================================================================

    def push_conversion(self, data: Dict[str, Any]) -> Optional[str]:
        """
        Push conversion data to MongoDB.

        Called automatically on job completion.

        Args:
            data: Conversion data dictionary

        Returns:
            Inserted document ID, or None if failed
        """
        if not self.ensure_connected():
            logger.warning("Cannot push conversion: MongoDB not connected")
            return None

        try:
            # Normalize data
            doc = self._normalize_conversion_data(data)
            doc["updated_at"] = datetime.utcnow()

            # Upsert by job_id
            result = self._collection.update_one(
                {"job_id": doc["job_id"]},
                {"$set": doc},
                upsert=True
            )

            if result.upserted_id:
                logger.info(f"Inserted new conversion: {doc['job_id']}")
                return str(result.upserted_id)
            else:
                logger.info(f"Updated existing conversion: {doc['job_id']}")
                return doc["job_id"]

        except Exception as e:
            logger.error(f"Failed to push conversion to MongoDB: {e}")
            return None

    def _normalize_conversion_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Normalize conversion data for MongoDB storage."""
        # Handle both Excel tracker format and API job format
        normalized = {
            "job_id": data.get("job_id") or data.get("filename", "unknown"),
            "filename": data.get("filename", ""),
            "isbn": data.get("isbn"),
            "title": data.get("title"),
            "publisher": data.get("publisher"),
            "authors": data.get("authors", []),
            "start_time": self._parse_datetime(data.get("start_time")),
            "end_time": self._parse_datetime(data.get("end_time")),
            "duration_seconds": data.get("duration_seconds"),
            "status": self._normalize_status(data.get("status", "in_progress")),
            "progress_percent": data.get("progress_percent", 0),
            "error_message": data.get("error_message"),
            "conversion_type": data.get("conversion_type", "PDF"),
            "template_type": data.get("template_type", "unknown"),
            "num_chapters": data.get("num_chapters", 0),
            "num_pages": data.get("num_pages", 0),
            "num_vector_images": data.get("num_vector_images", 0),
            "num_raster_images": data.get("num_raster_images", 0),
            "num_total_images": data.get("num_total_images", 0),
            "num_tables": data.get("num_tables", 0),
            "num_equations": data.get("num_equations", 0),
            "output_path": data.get("output_path"),
            "output_size_mb": data.get("output_size_mb"),
            "docx_path": data.get("docx_path"),
            "package_path": data.get("package_path"),
            "created_at": self._parse_datetime(data.get("created_at")) or datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        }

        # Calculate total images if not provided
        if normalized["num_total_images"] == 0:
            normalized["num_total_images"] = (
                normalized["num_vector_images"] + normalized["num_raster_images"]
            )

        # Calculate duration if not provided
        if normalized["duration_seconds"] is None and normalized["start_time"] and normalized["end_time"]:
            normalized["duration_seconds"] = (
                normalized["end_time"] - normalized["start_time"]
            ).total_seconds()

        return normalized

    def _parse_datetime(self, value) -> Optional[datetime]:
        """Parse datetime from various formats."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            try:
                # ISO format
                return datetime.fromisoformat(value.replace("Z", "+00:00"))
            except ValueError:
                try:
                    # Common format from Excel tracker
                    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    return None
        return None

    def _normalize_status(self, status) -> str:
        """Normalize status to consistent values."""
        if hasattr(status, 'value'):
            status = status.value
        status = str(status).lower().replace(" ", "_")
        # Map various status values
        status_map = {
            "in_progress": "in_progress",
            "success": "success",
            "completed": "success",
            "failure": "failure",
            "failed": "failure",
            "partial": "partial",
            "partial_success": "partial",
            "pending": "pending",
            "processing": "processing",
            "ready_for_review": "ready_for_review",
            "editing": "editing",
            "finalizing": "finalizing",
        }
        return status_map.get(status, status)

    def get_conversion(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get a specific conversion by job ID (which is the ISBN)."""
        if not self.ensure_connected():
            return None

        try:
            doc = self._collection.find_one({"job_id": job_id})
            if doc:
                doc["_id"] = str(doc["_id"])
            return doc
        except Exception as e:
            logger.error(f"Failed to get conversion {job_id}: {e}")
            return None

    def get_conversion_by_isbn(self, isbn: str) -> Optional[Dict[str, Any]]:
        """
        Get a specific conversion by ISBN.

        Since job_id is now the ISBN, this is equivalent to get_conversion,
        but provides a more semantic method name for ISBN-based lookups.

        Args:
            isbn: The ISBN number (same as job_id)

        Returns:
            Conversion document if found, None otherwise
        """
        if not self.ensure_connected():
            return None

        try:
            # Try by isbn field first, then by job_id (which should be the same)
            doc = self._collection.find_one({"$or": [{"isbn": isbn}, {"job_id": isbn}]})
            if doc:
                doc["_id"] = str(doc["_id"])
            return doc
        except Exception as e:
            logger.error(f"Failed to get conversion by ISBN {isbn}: {e}")
            return None

    def list_conversions(
        self,
        status: Optional[str] = None,
        limit: int = 50,
        skip: int = 0,
        sort_by: str = "created_at",
        sort_order: int = -1,
    ) -> List[Dict[str, Any]]:
        """
        List conversions with optional filtering.

        Args:
            status: Filter by status (optional)
            limit: Maximum number of results
            skip: Number of results to skip (for pagination)
            sort_by: Field to sort by
            sort_order: 1 for ascending, -1 for descending

        Returns:
            List of conversion documents
        """
        if not self.ensure_connected():
            return []

        try:
            query = {}
            if status:
                query["status"] = self._normalize_status(status)

            cursor = (
                self._collection.find(query)
                .sort(sort_by, sort_order)
                .skip(skip)
                .limit(limit)
            )

            results = []
            for doc in cursor:
                doc["_id"] = str(doc["_id"])
                results.append(doc)

            return results

        except Exception as e:
            logger.error(f"Failed to list conversions: {e}")
            return []

    # ========================================================================
    # DASHBOARD & ANALYTICS
    # ========================================================================

    def get_dashboard_stats(self) -> Dict[str, Any]:
        """
        Get dashboard statistics for UI display.

        Returns:
            Dictionary with dashboard statistics
        """
        if not self.ensure_connected():
            return self._empty_dashboard_stats()

        try:
            pipeline = [
                {
                    "$facet": {
                        "totals": [
                            {
                                "$group": {
                                    "_id": None,
                                    "total_conversions": {"$sum": 1},
                                    "total_pages": {"$sum": "$num_pages"},
                                    "total_images": {"$sum": "$num_total_images"},
                                    "total_tables": {"$sum": "$num_tables"},
                                    "total_chapters": {"$sum": "$num_chapters"},
                                    "avg_duration": {"$avg": "$duration_seconds"},
                                }
                            }
                        ],
                        "by_status": [
                            {"$group": {"_id": "$status", "count": {"$sum": 1}}}
                        ],
                        "by_type": [
                            {"$group": {"_id": "$conversion_type", "count": {"$sum": 1}}}
                        ],
                        "recent": [
                            {"$sort": {"created_at": -1}},
                            {"$limit": 10},
                            {
                                "$project": {
                                    "_id": 0,
                                    "job_id": 1,
                                    "filename": 1,
                                    "status": 1,
                                    "created_at": 1,
                                    "duration_seconds": 1,
                                    "num_pages": 1,
                                }
                            },
                        ],
                        "today": [
                            {
                                "$match": {
                                    "created_at": {
                                        "$gte": datetime.utcnow().replace(
                                            hour=0, minute=0, second=0, microsecond=0
                                        )
                                    }
                                }
                            },
                            {"$count": "count"},
                        ],
                        "this_week": [
                            {
                                "$match": {
                                    "created_at": {
                                        "$gte": datetime.utcnow() - timedelta(days=7)
                                    }
                                }
                            },
                            {"$count": "count"},
                        ],
                    }
                }
            ]

            result = list(self._collection.aggregate(pipeline))

            if not result:
                return self._empty_dashboard_stats()

            data = result[0]
            totals = data["totals"][0] if data["totals"] else {}

            # Build status counts
            status_counts = {item["_id"]: item["count"] for item in data["by_status"]}
            type_counts = {item["_id"]: item["count"] for item in data["by_type"]}

            return {
                "total_conversions": totals.get("total_conversions", 0),
                "successful": status_counts.get("success", 0),
                "failed": status_counts.get("failure", 0),
                "in_progress": status_counts.get("in_progress", 0) + status_counts.get("processing", 0),
                "pending": status_counts.get("pending", 0),
                "ready_for_review": status_counts.get("ready_for_review", 0),
                "total_pages_processed": totals.get("total_pages", 0),
                "total_images_extracted": totals.get("total_images", 0),
                "total_tables_extracted": totals.get("total_tables", 0),
                "total_chapters": totals.get("total_chapters", 0),
                "average_duration_seconds": round(totals.get("avg_duration", 0) or 0, 2),
                "conversions_today": data["today"][0]["count"] if data["today"] else 0,
                "conversions_this_week": data["this_week"][0]["count"] if data["this_week"] else 0,
                "by_status": status_counts,
                "by_type": type_counts,
                "recent_conversions": data["recent"],
            }

        except Exception as e:
            logger.error(f"Failed to get dashboard stats: {e}")
            return self._empty_dashboard_stats()

    def _empty_dashboard_stats(self) -> Dict[str, Any]:
        """Return empty dashboard stats structure."""
        return {
            "total_conversions": 0,
            "successful": 0,
            "failed": 0,
            "in_progress": 0,
            "pending": 0,
            "ready_for_review": 0,
            "total_pages_processed": 0,
            "total_images_extracted": 0,
            "total_tables_extracted": 0,
            "total_chapters": 0,
            "average_duration_seconds": 0,
            "conversions_today": 0,
            "conversions_this_week": 0,
            "by_status": {},
            "by_type": {},
            "recent_conversions": [],
        }

    def get_daily_stats(self, days: int = 30) -> List[Dict[str, Any]]:
        """
        Get daily conversion statistics for charts.

        Args:
            days: Number of days to include

        Returns:
            List of daily statistics
        """
        if not self.ensure_connected():
            return []

        try:
            start_date = datetime.utcnow() - timedelta(days=days)

            pipeline = [
                {"$match": {"created_at": {"$gte": start_date}}},
                {
                    "$group": {
                        "_id": {
                            "$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}
                        },
                        "total": {"$sum": 1},
                        "successful": {
                            "$sum": {"$cond": [{"$eq": ["$status", "success"]}, 1, 0]}
                        },
                        "failed": {
                            "$sum": {"$cond": [{"$eq": ["$status", "failure"]}, 1, 0]}
                        },
                        "pages": {"$sum": "$num_pages"},
                        "images": {"$sum": "$num_total_images"},
                    }
                },
                {"$sort": {"_id": 1}},
            ]

            return list(self._collection.aggregate(pipeline))

        except Exception as e:
            logger.error(f"Failed to get daily stats: {e}")
            return []

    def get_publisher_stats(self) -> List[Dict[str, Any]]:
        """Get statistics grouped by publisher."""
        if not self.ensure_connected():
            return []

        try:
            pipeline = [
                {"$match": {"publisher": {"$ne": None, "$ne": ""}}},
                {
                    "$group": {
                        "_id": "$publisher",
                        "total": {"$sum": 1},
                        "successful": {
                            "$sum": {"$cond": [{"$eq": ["$status", "success"]}, 1, 0]}
                        },
                        "failed": {
                            "$sum": {"$cond": [{"$eq": ["$status", "failure"]}, 1, 0]}
                        },
                        "total_pages": {"$sum": "$num_pages"},
                        "avg_duration": {"$avg": "$duration_seconds"},
                    }
                },
                {"$sort": {"total": -1}},
                {"$limit": 20},
            ]

            return list(self._collection.aggregate(pipeline))

        except Exception as e:
            logger.error(f"Failed to get publisher stats: {e}")
            return []


# ============================================================================
# SINGLETON INSTANCE
# ============================================================================

_mongodb_store: Optional[MongoDBStore] = None


def get_mongodb_store() -> MongoDBStore:
    """Get or create the singleton MongoDB store instance."""
    global _mongodb_store
    if _mongodb_store is None:
        _mongodb_store = MongoDBStore()
    return _mongodb_store


def init_mongodb(config: Optional[MongoDBConfig] = None) -> bool:
    """
    Initialize MongoDB connection.

    Call this at application startup.

    Args:
        config: Optional MongoDB configuration

    Returns:
        True if connected successfully
    """
    store = get_mongodb_store()
    if config:
        store.config = config
    return store.connect()


# ============================================================================
# EXCEL TO MONGODB SYNC
# ============================================================================

def sync_excel_to_mongodb(excel_path: Path) -> int:
    """
    Sync all data from Excel dashboard to MongoDB.

    Useful for initial migration or recovery.

    Args:
        excel_path: Path to conversion_dashboard.xlsx

    Returns:
        Number of records synced
    """
    if not MONGODB_AVAILABLE:
        logger.error("pymongo not installed")
        return 0

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        return 0

    store = get_mongodb_store()
    if not store.ensure_connected():
        logger.error("Cannot connect to MongoDB")
        return 0

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Conversions"]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        count = 0
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row_idx, col_idx).value
                if header:
                    # Map Excel headers to MongoDB fields
                    field_name = header.lower().replace(" ", "_").replace("#_", "num_").replace("%", "_percent")
                    row_data[field_name] = value

            if row_data.get("filename"):
                # Create job_id from filename + start_time
                job_id = f"{row_data.get('filename')}_{row_data.get('start_time', '')}"
                row_data["job_id"] = job_id

                store.push_conversion(row_data)
                count += 1

        logger.info(f"Synced {count} records from Excel to MongoDB")
        return count

    except Exception as e:
        logger.error(f"Failed to sync Excel to MongoDB: {e}")
        return 0
